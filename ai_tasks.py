"""AI augmentation tasks for the Startup Database.

Two tasks:
  daily_augment()   — pick 2 existing startups, find 2 similar new ones each, add to database
  monthly_update()  — for active startups, search for recent news and add as notes

Requires: ANTHROPIC_API_KEY environment variable and `pip install anthropic`
"""

import os, json, random, re, logging, time
import openpyxl
from datetime import datetime

logger = logging.getLogger(__name__)

BASE = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR     = os.environ.get('DATA_DIR', BASE)
EXCEL_DEFAULT = os.path.join(_DATA_DIR, 'Startup database.xlsx')
AI_LOG_FILE   = os.path.join(_DATA_DIR, 'ai_log.json')

ACTIVE_STATUSES = {'Early Exploration', 'Advanced discussions', 'Live collaboration'}

WEB_SEARCH_TOOL = {
    "type": "web_search_20250305",
    "name": "web_search",
    "max_uses": 5,
}

# Lighter version for monthly updates — fewer searches = far fewer input tokens
WEB_SEARCH_TOOL_LIGHT = {
    "type": "web_search_20250305",
    "name": "web_search",
    "max_uses": 2,
}

# ── Internal helpers ─────────────────────────────────────────────────────────

def _get_client():
    try:
        import anthropic
    except ImportError:
        raise RuntimeError('anthropic package not installed. Add it to requirements.txt and redeploy.')
    key = os.environ.get('ANTHROPIC_API_KEY')
    if not key:
        raise ValueError('ANTHROPIC_API_KEY environment variable is not set.')
    return anthropic.Anthropic(api_key=key)


def _build_hmap(ws):
    hmap = {}
    for c in range(1, 120):
        v = ws.cell(row=1, column=c).value
        if v:
            hmap[str(v).strip().lstrip('\ufeff')] = c
    return hmap


def _read_startups(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    hmap = _build_hmap(ws)
    name_col = hmap.get('Company name', 1)
    startups = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col).value
        if not name or str(name).strip().lower() in ('nan', 'none', ''):
            continue
        row_data = {h: ws.cell(row=r, column=c).value for h, c in hmap.items()}
        startups.append(row_data)
    return startups


def _find_company_row(ws, name_col, name):
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        if v and str(v).strip().lower() == name.lower():
            return r
    return None


def _write_new_startup(ws, hmap, name, website=None, country=None,
                       description=None, category=None, added_by='AI (daily augment)'):
    """Append a new startup row."""
    next_row = ws.max_row + 1
    def w(header, value):
        col = hmap.get(header)
        if col and value:
            ws.cell(row=next_row, column=col).value = value
    w('Company name', name)
    w('WBWSite',      website)
    w('Country',      country)
    w('Description',  description)
    w('Category',     category)
    w('Status',       'No contact')
    w('Added by',     added_by)


def _add_note(ws, hmap, company_row, text, date=None):
    """Append a note to an existing startup row."""
    if date is None:
        date = datetime.now().strftime('%Y-%m-%d')
    note_cols = sorted(
        c for h, c in hmap.items()
        if h == 'Notes' or (h.startswith('Notes') and h[5:].strip().lstrip('.').isdigit())
    )
    filled = sum(1 for c in note_cols if ws.cell(row=company_row, column=c).value is not None)
    n = filled + 1
    note_header = 'Notes' if n == 1 else f'Notes {n}'
    col = hmap.get(note_header)
    if col is None:
        col = max(hmap.values()) + 1
        ws.cell(row=1, column=col).value = note_header
        hmap[note_header] = col
    ws.cell(row=company_row, column=col).value = f"{date}: {text}"


def _call_claude(client, prompt, max_tokens=1024, search_tool=None):
    """Call Claude with web search, return concatenated text from all text blocks.
    Retries up to 3 times with exponential backoff on rate-limit (429) errors.
    """
    if search_tool is None:
        search_tool = WEB_SEARCH_TOOL
    for attempt in range(3):
        try:
            response = client.messages.create(
                model='claude-sonnet-4-6',
                max_tokens=max_tokens,
                tools=[search_tool],
                messages=[{'role': 'user', 'content': prompt}],
            )
            return ''.join(b.text for b in response.content if b.type == 'text')
        except Exception as e:
            err = str(e)
            if ('429' in err or 'rate_limit' in err.lower()) and attempt < 2:
                wait = 65 * (attempt + 1)   # 65 s, then 130 s
                logger.warning(f'Rate limited — waiting {wait}s before retry {attempt + 1}')
                time.sleep(wait)
            else:
                raise


def _parse_json(text):
    """Extract a JSON array or object from free text."""
    # Try ```json ... ``` fenced block first
    m = re.search(r'```(?:json)?\s*([\[{].*?[\]}])\s*```', text, re.DOTALL)
    if m:
        return json.loads(m.group(1))
    # Try bare JSON array or object
    m = re.search(r'(\[[\s\S]*\]|\{[\s\S]*\})', text)
    if m:
        return json.loads(m.group(1))
    raise ValueError(f'No JSON found in:\n{text[:300]}')


def _append_log(entry):
    log = []
    if os.path.exists(AI_LOG_FILE):
        try:
            with open(AI_LOG_FILE) as f:
                log = json.load(f)
        except Exception:
            pass
    log.insert(0, entry)
    with open(AI_LOG_FILE, 'w') as f:
        json.dump(log[:100], f, indent=2)


# ── Public task functions ────────────────────────────────────────────────────

def daily_augment(excel_path=None):
    """Pick 2 seed startups, find 2 similar real ones each, add to database.

    Returns a dict with keys: type, timestamp, seeds, added, errors.
    """
    if excel_path is None:
        excel_path = EXCEL_DEFAULT

    client = _get_client()
    existing = _read_startups(excel_path)
    existing_names_lc = {str(s.get('Company name', '')).strip().lower() for s in existing}

    # Prefer seeds with descriptions for better Claude context
    seeds_pool = [s for s in existing if s.get('Description') and s.get('Company name')]
    if not seeds_pool:
        seeds_pool = existing
    seeds = random.sample(seeds_pool, min(2, len(seeds_pool)))

    added, errors = [], []

    for seed in seeds:
        seed_name = str(seed.get('Company name', '')).strip()
        seed_cat  = seed.get('Category', '') or ''
        seed_desc = str(seed.get('Description', '') or '')[:300]
        seed_cty  = seed.get('Country', '') or ''

        prompt = f"""I maintain a database of energy and cleantech startups for EDF Energy UK.

One startup already in my database:
- Name: {seed_name}
- Category: {seed_cat}
- Country: {seed_cty}
- Description: {seed_desc}

Search the web and identify exactly 2 REAL, currently-active UK-based startups that are similar to {seed_name}
in technology or business model. Requirements:
• Must be based in the United Kingdom (England, Scotland, Wales, or Northern Ireland)
• Must exist and be operational
• Must be in the energy, cleantech, sustainability, nuclear, or adjacent sector
• Must NOT be {seed_name} itself
• Must NOT be a large incumbent (no BP, Shell, EDF, National Grid, etc.) — startups only (founded after 2005, fewer than 500 employees)

Return ONLY a JSON array (no other text) with exactly 2 objects:
[
  {{
    "name": "Exact company name",
    "website": "https://...",
    "country": "United Kingdom",
    "description": "2-3 sentences describing what they do and why they are similar to {seed_name}",
    "category": "most relevant sub-category (e.g. Solar PV, Battery Storage & BESS, Heat Pumps (Residential), Hydrogen Production, AI & Data Analytics, Nuclear Operations & Digital Tools, etc.)"
  }},
  {{ ... }}
]"""

        try:
            text = _call_claude(client, prompt, max_tokens=1500)
            companies = _parse_json(text)
            if isinstance(companies, dict):
                companies = [companies]

            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            hmap = _build_hmap(ws)
            changed = False

            for co in companies[:2]:
                name = str(co.get('name', '') or '').strip()
                if not name or len(name) < 3 or name.lower() in ('none', 'n/a', 'unknown', ''):
                    continue
                if name.lower() in existing_names_lc:
                    continue
                _write_new_startup(
                    ws, hmap,
                    name=name,
                    website=co.get('website'),
                    country=co.get('country'),
                    description=co.get('description'),
                    category=co.get('category'),
                    added_by=f'AI (similar to {seed_name})',
                )
                existing_names_lc.add(name.lower())
                added.append({'name': name, 'similar_to': seed_name})
                changed = True
                logger.info(f'daily_augment: added {name}')

            if changed:
                wb.save(excel_path)

        except Exception as e:
            msg = f'Seed "{seed_name}": {e}'
            errors.append(msg)
            logger.error(f'daily_augment error — {msg}')

    result = {
        'type': 'daily_augment',
        'timestamp': datetime.now().isoformat(),
        'seeds': [str(s.get('Company name', '')) for s in seeds],
        'added': added,
        'errors': errors,
    }
    _append_log(result)
    return result


def monthly_update(excel_path=None):
    """Search for recent news on active startups and add as notes.

    Processes startups in Early Exploration, Advanced discussions, or Live collaboration.
    Caps at 10 per run to control API costs.

    Returns a dict with keys: type, timestamp, checked, updated, errors.
    """
    if excel_path is None:
        excel_path = EXCEL_DEFAULT

    client = _get_client()
    existing = _read_startups(excel_path)

    active = [s for s in existing if s.get('Status') == 'Advanced discussions' and s.get('Company name')]
    if not active:
        result = {
            'type': 'monthly_update',
            'timestamp': datetime.now().isoformat(),
            'checked': 0,
            'updated': [],
            'errors': ['No startups found with status "Advanced discussions".'],
        }
        _append_log(result)
        return result

    # Process all Advanced discussions startups (cap at 30 to control API costs)
    targets = sorted(active, key=lambda s: s.get('Company name', ''))[:30]

    updated, errors = [], []
    today = datetime.now().strftime('%Y-%m-%d')

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    hmap = _build_hmap(ws)
    name_col = hmap.get('Company name', 1)
    changed = False

    for startup in targets:
        name   = str(startup.get('Company name', '')).strip()
        desc   = str(startup.get('Description', '') or '')[:250]
        status = startup.get('Status', '')
        cat    = startup.get('Category', '') or ''

        prompt = f"""I'm evaluating startups for potential partnership at EDF Energy.

Company: {name}
Sector: {cat}
Current relationship status: {status}
{'Description: ' + desc if desc else ''}

Please search the web for significant news about {name} from the last 6 months. Focus on:
- Funding rounds or acquisitions
- New major customers or partnerships
- Product launches or technology breakthroughs
- Executive changes
- Any red flags (financial trouble, regulatory issues, pivots)

Write a concise summary (3–5 sentences max) of the most important updates.
If you find no meaningful recent news, respond with exactly: NO_NEWS"""

        try:
            text = _call_claude(client, prompt, max_tokens=512,
                                search_tool=WEB_SEARCH_TOOL_LIGHT).strip()
            if not text or text.upper().startswith('NO_NEWS'):
                time.sleep(20)   # pace between calls even when no update
                continue

            row = _find_company_row(ws, name_col, name)
            if not row:
                continue

            note = f"[AI news update] {text}"
            _add_note(ws, hmap, row, note, date=today)
            updated.append({
                'name': name,
                'summary': text[:120] + ('…' if len(text) > 120 else ''),
            })
            changed = True
            logger.info(f'monthly_update: added note for {name}')

        except Exception as e:
            msg = f'{name}: {e}'
            errors.append(msg)
            logger.error(f'monthly_update error — {msg}')

        time.sleep(20)  # ~20 s gap: ≤3 calls/min × ~5k tokens = ~15k tokens/min (well under 30k limit)

    if changed:
        wb.save(excel_path)

    result = {
        'type': 'monthly_update',
        'timestamp': datetime.now().isoformat(),
        'checked': len(targets),
        'updated': updated,
        'errors': errors,
    }
    _append_log(result)
    return result


def find_similar_for_startup(startup_name, excel_path=None):
    """Find 3 similar real startups for a specific startup and add them to the database.

    Returns a dict with keys: type, timestamp, seed, added, errors.
    """
    if excel_path is None:
        excel_path = EXCEL_DEFAULT

    client = _get_client()
    existing = _read_startups(excel_path)
    existing_names_lc = {str(s.get('Company name', '')).strip().lower() for s in existing}

    seed = next(
        (s for s in existing if str(s.get('Company name', '')).strip().lower() == startup_name.lower()),
        None
    )
    if not seed:
        raise ValueError(f'Startup "{startup_name}" not found in database')

    seed_name = str(seed.get('Company name', '')).strip()
    seed_cat  = seed.get('Category', '') or ''
    seed_desc = str(seed.get('Description', '') or '')[:300]
    seed_cty  = seed.get('Country', '') or ''

    prompt = f"""I maintain a database of energy and cleantech startups for EDF Energy UK.

One startup already in my database:
- Name: {seed_name}
- Category: {seed_cat}
- Country: {seed_cty}
- Description: {seed_desc}

Search the web and identify exactly 2 REAL, currently-active UK-based startups that are similar to {seed_name}
in technology or business model. Requirements:
• Must be based in the United Kingdom (England, Scotland, Wales, or Northern Ireland)
• Must actually exist and be operational
• Must be in the energy, cleantech, sustainability, nuclear, or adjacent sector
• Must NOT be {seed_name} itself
• Must NOT be large incumbents (no BP, Shell, National Grid, etc.) — startups only (founded after 2005, fewer than 500 employees)

Return ONLY a JSON array with exactly 2 objects:
[
  {{
    "name": "Exact company name",
    "website": "https://...",
    "country": "United Kingdom",
    "description": "2-3 sentences describing what they do and why they are similar to {seed_name}",
    "category": "{seed_cat if seed_cat else 'most relevant sub-category'}"
  }},
  {{ ... }}
]"""

    errors = []
    try:
        text = _call_claude(client, prompt, max_tokens=2000)
        companies = _parse_json(text)
        if isinstance(companies, dict):
            companies = [companies]
    except Exception as e:
        errors.append(str(e))
        result = {'type': 'find_similar', 'timestamp': datetime.now().isoformat(),
                  'seed': seed_name, 'added': [], 'errors': errors}
        _append_log(result)
        return result

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    hmap = _build_hmap(ws)
    added = []

    for co in companies[:2]:
        name = str(co.get('name', '') or '').strip()
        if not name or len(name) < 3 or name.lower() in ('none', 'n/a', 'unknown', ''):
            continue
        if name.lower() in existing_names_lc:
            continue
        _write_new_startup(
            ws, hmap,
            name=name,
            website=co.get('website'),
            country=co.get('country'),
            description=co.get('description'),
            category=co.get('category') or seed_cat,
            added_by=f'AI (similar to {seed_name})',
        )
        existing_names_lc.add(name.lower())
        added.append({'name': name, 'similar_to': seed_name})
        logger.info(f'find_similar: added {name} (similar to {seed_name})')

    if added:
        wb.save(excel_path)

    result = {
        'type': 'find_similar',
        'timestamp': datetime.now().isoformat(),
        'seed': seed_name,
        'added': added,
        'errors': errors,
    }
    _append_log(result)
    return result


def fill_missing_websites(excel_path=None):
    """For every startup missing a website (or with only LinkedIn/Crunchbase),
    search the web and write the official URL back to the spreadsheet.

    Returns a dict with keys: type, timestamp, checked, updated, errors.
    """
    if excel_path is None:
        excel_path = EXCEL_DEFAULT

    def _needs_website(raw):
        if not raw or str(raw).strip().lower() in ('', '-', 'nan', 'none'):
            return True
        u = str(raw).strip().lower()
        return 'linkedin.com' in u or 'crunchbase.com' in u

    client  = _get_client()
    existing = _read_startups(excel_path)
    targets  = [s for s in existing
                if s.get('Company name') and _needs_website(s.get('WBWSite'))]

    if not targets:
        result = {
            'type': 'fill_websites', 'timestamp': datetime.now().isoformat(),
            'checked': 0, 'updated': [], 'errors': [],
        }
        _append_log(result)
        return result

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    hmap = _build_hmap(ws)
    name_col    = hmap.get('Company name', 1)
    website_col = hmap.get('WBWSite')
    if not website_col:
        raise ValueError('WBWSite column not found in spreadsheet')

    updated, errors = [], []
    changed = False

    for startup in targets:
        name = str(startup.get('Company name', '')).strip()
        desc = str(startup.get('Description', '') or '')[:200]
        cat  = startup.get('Category', '') or ''

        prompt = (
            f'Find the official company website URL for "{name}", '
            f'a startup in the {cat or "energy/cleantech"} sector.\n'
            + (f'Description: {desc}\n' if desc else '')
            + 'Return ONLY the URL (e.g. https://example.com). '
            'Do not include LinkedIn, Crunchbase, or news articles. '
            'If no official website can be found, reply: NOT_FOUND'
        )

        try:
            text = _call_claude(client, prompt, max_tokens=80,
                                search_tool=WEB_SEARCH_TOOL_LIGHT).strip()
            # Extract first token — should be a URL
            url = text.split()[0] if text else ''
            url = url.rstrip('.,;)')
            if (not url or 'NOT_FOUND' in url.upper()
                    or not url.startswith('http')
                    or 'linkedin.com' in url.lower()
                    or 'crunchbase.com' in url.lower()):
                time.sleep(5)
                continue

            row = _find_company_row(ws, name_col, name)
            if not row:
                time.sleep(5)
                continue

            ws.cell(row=row, column=website_col).value = url
            updated.append({'name': name, 'website': url})
            changed = True
            logger.info(f'fill_websites: {name} → {url}')

        except Exception as e:
            msg = f'{name}: {e}'
            errors.append(msg)
            logger.error(f'fill_websites error — {msg}')

        time.sleep(5)

    if changed:
        wb.save(excel_path)

    result = {
        'type': 'fill_websites',
        'timestamp': datetime.now().isoformat(),
        'checked': len(targets),
        'updated': updated,
        'errors': errors,
    }
    _append_log(result)
    return result


def fill_missing_descriptions(excel_path=None):
    """For every startup missing a description (or with a very short one),
    search the web and write a 2-3 sentence description back to the spreadsheet.

    Returns a dict with keys: type, timestamp, checked, updated, errors.
    """
    if excel_path is None:
        excel_path = EXCEL_DEFAULT

    def _needs_description(val):
        if not val or str(val).strip().lower() in ('', '-', 'nan', 'none'):
            return True
        return len(str(val).strip()) < 30   # too short to be useful

    client   = _get_client()
    existing = _read_startups(excel_path)
    targets  = [s for s in existing
                if s.get('Company name') and _needs_description(s.get('Description'))]

    if not targets:
        result = {
            'type': 'fill_descriptions', 'timestamp': datetime.now().isoformat(),
            'checked': 0, 'updated': [], 'errors': [],
        }
        _append_log(result)
        return result

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    hmap = _build_hmap(ws)
    name_col = hmap.get('Company name', 1)
    desc_col = hmap.get('Description')
    if not desc_col:
        raise ValueError('Description column not found in spreadsheet')

    updated, errors = [], []
    changed = False

    for startup in targets:
        name    = str(startup.get('Company name', '')).strip()
        website = str(startup.get('WBWSite', '') or '').strip()
        cat     = startup.get('Category', '') or ''

        prompt = (
            f'Search the web for information about "{name}", '
            f'a startup in the {cat or "energy/cleantech"} sector'
            + (f' ({website})' if website and website.startswith('http') else '')
            + '.\n'
            'Write a factual 2-3 sentence description of what the company does, '
            'its core technology or product, and who it serves. '
            'Be specific and concise. Do not use marketing language. '
            'If you cannot find reliable information about this company, reply: NOT_FOUND'
        )

        try:
            text = _call_claude(client, prompt, max_tokens=200,
                                search_tool=WEB_SEARCH_TOOL_LIGHT).strip()

            if not text or 'NOT_FOUND' in text.upper()[:20]:
                time.sleep(5)
                continue

            # Truncate to a reasonable length
            if len(text) > 500:
                text = text[:500].rsplit('.', 1)[0] + '.'

            row = _find_company_row(ws, name_col, name)
            if not row:
                time.sleep(5)
                continue

            ws.cell(row=row, column=desc_col).value = text
            updated.append({'name': name})
            changed = True
            logger.info(f'fill_descriptions: updated {name}')

        except Exception as e:
            msg = f'{name}: {e}'
            errors.append(msg)
            logger.error(f'fill_descriptions error — {msg}')

        time.sleep(5)

    if changed:
        wb.save(excel_path)

    result = {
        'type': 'fill_descriptions',
        'timestamp': datetime.now().isoformat(),
        'checked': len(targets),
        'updated': updated,
        'errors': errors,
    }
    _append_log(result)
    return result


def get_log(n=20):
    """Return the n most recent AI task log entries."""
    if not os.path.exists(AI_LOG_FILE):
        return []
    try:
        with open(AI_LOG_FILE) as f:
            return json.load(f)[:n]
    except Exception:
        return []
