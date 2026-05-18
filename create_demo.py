"""
Generate demo-database.xlsx — a realistic dummy dataset for Startup Scout.

Run:  python3 create_demo.py
Output: demo-database.xlsx  (ready to upload on the setup page)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import os, random

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'demo-database.xlsx')

# ── Column definitions (must match app expectations) ─────────────────────────
BU_COLS = [
    'Nuclear operations', 'Nuclear Services', 'Dalkia', 'BWS', 'Retail',
    'EDF Heat Pumps', 'Contact Solar', 'Hynamics', 'R&D', 'EDF Power Solutions', 'Pod',
]

HEADERS = (
    ['Company name'] + BU_COLS +
    ['Category', 'Status', 'Country', 'Description', 'Contact',
     'WBWSite', 'Added by', 'EDF PV  portfolio ', 'NDA signed',
     'Key contact', 'Key contact email', 'Relationship owner', 'Avoid',
     'Notes', 'Notes 2', 'Notes 3']
)

# ── Demo data ────────────────────────────────────────────────────────────────
STATUSES = [
    'No contact',
    'Met but no live discussions',
    'Early Exploration',
    'Advanced discussions',
    'Live collaboration',
]

STARTUPS = [
    # ── Live collaboration ────────────────────────────────────────────────
    {
        'Company name': 'Voltora Energy',
        'Category': 'Battery Storage & BESS',
        'Status': 'Live collaboration',
        'Country': 'United Kingdom',
        'Description': 'Voltora develops second-life battery systems for grid-scale storage, repurposing retired EV battery packs into 4-hour BESS units at 60% of the cost of new cells.',
        'WBWSite': 'https://voltora.energy',
        'Added by': 'Sarah Mitchell',
        'EDF PV  portfolio ': 'Yes',
        'NDA signed': 'Yes',
        'Key contact': 'James Hartley',
        'Key contact email': 'j.hartley@voltora.energy',
        'Relationship owner': 'Sarah Mitchell',
        'Avoid': 'No',
        'BUs': ['EDF Power Solutions', 'R&D'],
        'Notes': '2024-03-15: Signed MoU for pilot at Grain Power Station. 50 units to be deployed Q3.',
        'Notes 2': '2024-05-02: Pilot going well — 94% round-trip efficiency reported. Board presentation scheduled.',
        'Notes 3': '2024-06-18: Expanding pilot scope to include second site at West Burton.',
    },
    {
        'Company name': 'Heliogen Grid',
        'Category': 'Smart Grid & Flexibility',
        'Status': 'Live collaboration',
        'Country': 'United Kingdom',
        'Description': 'Heliogen Grid uses ML-based demand forecasting to optimise distribution network dispatch, reducing curtailment by up to 18% in trials with UK DNOs.',
        'WBWSite': 'https://heliogengrid.com',
        'Added by': 'Tom Whitfield',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'Yes',
        'Key contact': 'Priya Anand',
        'Key contact email': 'priya@heliogengrid.com',
        'Relationship owner': 'Tom Whitfield',
        'Avoid': 'No',
        'BUs': ['EDF Power Solutions', 'Retail'],
        'Notes': '2024-04-10: Joint paper published in IEEE Power journal. Positive coverage.',
        'Notes 2': '2024-06-01: Expanding into EDF metering data integration. Legal reviewing data-sharing agreement.',
    },

    # ── Advanced discussions ──────────────────────────────────────────────
    {
        'Company name': 'Thermify',
        'Category': 'Heat Pumps (Residential)',
        'Status': 'Advanced discussions',
        'Country': 'United Kingdom',
        'Description': 'Thermify installs and finances hybrid heat pump systems for social housing landlords, bundling installation, maintenance and a fixed-price energy tariff into a single monthly payment.',
        'WBWSite': 'https://thermify.co.uk',
        'Added by': 'Rachel Osei',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'Yes',
        'Key contact': 'Oliver Barnes',
        'Key contact email': 'oliver.barnes@thermify.co.uk',
        'Relationship owner': 'Rachel Osei',
        'Avoid': 'No',
        'BUs': ['EDF Heat Pumps', 'Retail'],
        'Notes': '2024-05-20: Proposal reviewed by EDF Heat Pumps team — positive. Commercial terms being drafted.',
        'Notes 2': '2024-07-03: Term sheet circulated. Legal on both sides reviewing.',
    },
    {
        'Company name': 'NuclearMind AI',
        'Category': 'Nuclear Operations & Digital Tools',
        'Status': 'Advanced discussions',
        'Country': 'United Kingdom',
        'Description': 'NuclearMind AI applies computer vision and NLP to automate safety documentation review and anomaly detection in nuclear plant operations, cutting review time by 70%.',
        'WBWSite': 'https://nuclearmind.ai',
        'Added by': 'David Ferguson',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'Yes',
        'Key contact': 'Dr. Fatima Chaudhry',
        'Key contact email': 'f.chaudhry@nuclearmind.ai',
        'Relationship owner': 'David Ferguson',
        'Avoid': 'No',
        'BUs': ['Nuclear operations', 'Nuclear Services', 'R&D'],
        'Notes': '2024-04-22: Demo at Hinkley Point B — very well received by ops team.',
        'Notes 2': '2024-06-14: Formal evaluation underway. Security clearance process started for their engineers.',
    },
    {
        'Company name': 'ClearWatt',
        'Category': 'AI & Data Analytics',
        'Status': 'Advanced discussions',
        'Country': 'United Kingdom',
        'Description': 'ClearWatt provides carbon-intensity-aware EV charging software, shifting charge schedules automatically to match grid greenness and halving the carbon footprint of fleet charging.',
        'WBWSite': 'https://clearwatt.co.uk',
        'Added by': 'Sarah Mitchell',
        'EDF PV  portfolio ': 'Yes',
        'NDA signed': 'No',
        'Key contact': 'Niamh Clarke',
        'Key contact email': 'niamh@clearwatt.co.uk',
        'Relationship owner': 'Sarah Mitchell',
        'Avoid': 'No',
        'BUs': ['Retail', 'EDF Power Solutions'],
        'Notes': '2024-03-08: Introduced by Innovate UK. Strong alignment with EDF fleet electrification programme.',
        'Notes 2': '2024-05-30: POC scoped for EDF van fleet at Exeter depot. Starting Sept.',
    },
    {
        'Company name': 'Carbonloop',
        'Category': 'Carbon Capture & Storage',
        'Status': 'Advanced discussions',
        'Country': 'United Kingdom',
        'Description': 'Carbonloop deploys modular direct-air capture units co-located with industrial sites, using low-grade waste heat to drive the capture process and cut energy costs by 40%.',
        'WBWSite': 'https://carbonloop.earth',
        'Added by': 'Tom Whitfield',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'Yes',
        'Key contact': 'Ben Okafor',
        'Key contact email': 'ben@carbonloop.earth',
        'Relationship owner': 'Tom Whitfield',
        'Avoid': 'No',
        'BUs': ['R&D', 'Dalkia'],
        'Notes': '2024-06-25: Site visit to West Burton completed. Feasibility study commissioned.',
    },

    # ── Early Exploration ──────────────────────────────────────────────────
    {
        'Company name': 'Aerobotics UK',
        'Category': 'AI & Data Analytics',
        'Status': 'Early Exploration',
        'Country': 'United Kingdom',
        'Description': 'Aerobotics uses drone-based thermal imaging and AI to detect faults in solar PV installations, reducing inspection costs by 80% versus manual surveys.',
        'WBWSite': 'https://aerobotics.com',
        'Added by': 'Rachel Osei',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Marcus Webb',
        'Key contact email': 'm.webb@aerobotics.com',
        'Relationship owner': 'Rachel Osei',
        'Avoid': 'No',
        'BUs': ['Contact Solar', 'R&D'],
        'Notes': '2024-06-10: Met at Solar UK conference. Interesting tech — scheduling follow-up call.',
    },
    {
        'Company name': 'Hydrolink Systems',
        'Category': 'Water Technology & Management',
        'Status': 'Early Exploration',
        'Country': 'United Kingdom',
        'Description': 'Hydrolink develops low-power IoT sensors for real-time leak detection in water distribution networks, achieving 95% detection accuracy with a 3-year battery life.',
        'WBWSite': 'https://hydrolinksystems.co.uk',
        'Added by': 'David Ferguson',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Alicia Drummond',
        'Key contact email': 'a.drummond@hydrolinksystems.co.uk',
        'Relationship owner': 'David Ferguson',
        'Avoid': 'No',
        'BUs': ['BWS'],
        'Notes': '2024-05-14: Intro call done. Good product — exploring fit with BWS network monitoring.',
    },
    {
        'Company name': 'Gridshift',
        'Category': 'Smart Grid & Flexibility',
        'Status': 'Early Exploration',
        'Country': 'United Kingdom',
        'Description': 'Gridshift aggregates domestic smart appliances — heat pumps, EV chargers and hot water cylinders — into a virtual power plant, trading flexibility in the Balancing Mechanism.',
        'WBWSite': 'https://gridshift.energy',
        'Added by': 'Tom Whitfield',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Kofi Asante',
        'Key contact email': 'kofi@gridshift.energy',
        'Relationship owner': 'Tom Whitfield',
        'Avoid': 'No',
        'BUs': ['Retail', 'EDF Power Solutions'],
        'Notes': '2024-07-01: Referred by National Grid ESO. Demo booked for end of month.',
    },
    {
        'Company name': 'SkyHydrogen',
        'Category': 'Hydrogen Production',
        'Status': 'Early Exploration',
        'Country': 'United Kingdom',
        'Description': 'SkyHydrogen produces green hydrogen via an electrolysis-as-a-service model, siting modular 1MW electrolyser units at renewable energy sites and selling H2 under long-term offtake agreements.',
        'WBWSite': 'https://skyhydrogen.co.uk',
        'Added by': 'Sarah Mitchell',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Elena Torres',
        'Key contact email': 'e.torres@skyhydrogen.co.uk',
        'Relationship owner': 'Sarah Mitchell',
        'Avoid': 'No',
        'BUs': ['Hynamics', 'R&D'],
        'Notes': '2024-06-05: Strong team. Exploring colocation with EDF wind assets.',
    },
    {
        'Company name': 'Ambri Storage',
        'Category': 'Battery Storage & BESS',
        'Status': 'Early Exploration',
        'Country': 'United States',
        'Description': 'Ambri makes liquid metal batteries designed for long-duration grid storage, offering 20-year calendar life and inherently safe chemistry with no fire risk.',
        'WBWSite': 'https://ambri.com',
        'Added by': 'David Ferguson',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Chris Park',
        'Key contact email': 'cpark@ambri.com',
        'Relationship owner': 'David Ferguson',
        'Avoid': 'No',
        'BUs': ['EDF Power Solutions'],
        'Notes': '2024-04-18: Met at Energy Storage Summit. US-based but looking at UK partnerships.',
    },
    {
        'Company name': 'Ferrobots',
        'Category': 'Nuclear Operations & Digital Tools',
        'Status': 'Early Exploration',
        'Country': 'United Kingdom',
        'Description': 'Ferrobots manufactures magnetic-crawling inspection robots for pressure vessel and piping inspection inside nuclear facilities, eliminating the need for human entry into radiation zones.',
        'WBWSite': 'https://ferrobots.co.uk',
        'Added by': 'David Ferguson',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Dr. Ian Pearce',
        'Key contact email': 'i.pearce@ferrobots.co.uk',
        'Relationship owner': 'David Ferguson',
        'Avoid': 'No',
        'BUs': ['Nuclear operations', 'Nuclear Services'],
        'Notes': '2024-07-10: Very early stage. Founder background from Rolls-Royce Nuclear. Watching closely.',
    },
    {
        'Company name': 'Lumenaza',
        'Category': 'Smart Grid & Flexibility',
        'Status': 'Early Exploration',
        'Country': 'Germany',
        'Description': 'Lumenaza provides white-label energy community software that enables utilities to create peer-to-peer local energy trading platforms for prosumers with rooftop solar.',
        'WBWSite': 'https://lumenaza.de',
        'Added by': 'Tom Whitfield',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Markus Bauer',
        'Key contact email': 'm.bauer@lumenaza.de',
        'Relationship owner': 'Tom Whitfield',
        'Avoid': 'No',
        'BUs': ['Retail'],
        'Notes': '2024-05-22: Interesting model for community solar. Not UK-native but could be adapted.',
    },
    {
        'Company name': 'Upside Energy',
        'Category': 'Smart Grid & Flexibility',
        'Status': 'Early Exploration',
        'Country': 'United Kingdom',
        'Description': 'Upside Energy aggregates behind-the-meter flexibility from commercial batteries and EV fleets, selling response services into National Grid ancillary markets.',
        'WBWSite': 'https://upsideenergy.co.uk',
        'Added by': 'Rachel Osei',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Relationship owner': 'Rachel Osei',
        'Avoid': 'No',
        'BUs': ['EDF Power Solutions', 'Retail'],
        'Notes': '2024-06-28: Good team. Assessing whether their flexibility stack complements or competes with ours.',
    },

    # ── Met but no live discussions ───────────────────────────────────────
    {
        'Company name': 'Biocraft Energy',
        'Category': 'Sustainable Fuels & Biomass',
        'Status': 'Met but no live discussions',
        'Country': 'United Kingdom',
        'Description': 'Biocraft converts agricultural residues into bioLPG and biomethane using a proprietary low-temperature pyrolysis process, targeting rural off-grid heating markets.',
        'WBWSite': 'https://biocraft.energy',
        'Added by': 'Sarah Mitchell',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Hannah Scott',
        'Key contact email': 'h.scott@biocraft.energy',
        'Relationship owner': 'Sarah Mitchell',
        'Avoid': 'No',
        'BUs': ['Dalkia'],
        'Notes': '2024-02-14: Met at COP28 side event. Interesting tech but limited synergy with current EDF priorities.',
    },
    {
        'Company name': 'Nuvve UK',
        'Category': 'EV Charging & V2G',
        'Status': 'Met but no live discussions',
        'Country': 'United Kingdom',
        'Description': 'Nuvve provides vehicle-to-grid (V2G) charging hardware and aggregation software, allowing EV batteries to export power back to the grid during peak demand periods.',
        'WBWSite': 'https://nuvve.com',
        'Added by': 'Tom Whitfield',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Sophie Laurent',
        'Key contact email': 's.laurent@nuvve.com',
        'Relationship owner': 'Tom Whitfield',
        'Avoid': 'No',
        'BUs': ['Retail'],
        'Notes': '2024-03-01: Good product but market timing uncertain. Revisit when V2G regs clearer.',
    },
    {
        'Company name': 'Kairos Power',
        'Category': 'Nuclear Operations & Digital Tools',
        'Status': 'Met but no live discussions',
        'Country': 'United States',
        'Description': 'Kairos Power is developing a fluoride-salt-cooled high-temperature reactor (KP-FHR) for industrial process heat and power, with a target overnight cost below £3,000/kW.',
        'WBWSite': 'https://kairospower.com',
        'Added by': 'David Ferguson',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Dr. Mike Laufer',
        'Key contact email': 'mlaufer@kairospower.com',
        'Relationship owner': 'David Ferguson',
        'Avoid': 'No',
        'BUs': ['Nuclear operations', 'R&D'],
        'Notes': '2024-01-20: Interesting SMR concept but very long timelines. Park for 2 years.',
    },
    {
        'Company name': 'PassivSystems',
        'Category': 'Heat Pumps (Residential)',
        'Status': 'Met but no live discussions',
        'Country': 'United Kingdom',
        'Description': 'PassivSystems optimises home heating schedules using weather forecasting and occupancy data, reducing gas and heat pump energy consumption by up to 25% with no hardware.',
        'WBWSite': 'https://passivsystems.com',
        'Added by': 'Rachel Osei',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Relationship owner': 'Rachel Osei',
        'Avoid': 'No',
        'BUs': ['EDF Heat Pumps', 'Retail'],
        'Notes': '2024-04-30: Good tech but CEO leaving. Put on hold until new leadership in place.',
    },
    {
        'Company name': 'Recycleye',
        'Category': 'Circular Economy & Waste',
        'Status': 'Met but no live discussions',
        'Country': 'United Kingdom',
        'Description': 'Recycleye uses computer vision and robotics to automate material sorting in waste processing facilities, increasing recycling rates from 60% to 90% in pilot sites.',
        'WBWSite': 'https://recycleye.com',
        'Added by': 'Tom Whitfield',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Victor Dewulf',
        'Key contact email': 'victor@recycleye.com',
        'Relationship owner': 'Tom Whitfield',
        'Avoid': 'No',
        'BUs': ['Dalkia'],
        'Notes': '2024-02-28: Niche fit with Dalkia waste ops. Dalkia team to follow up independently.',
    },
    {
        'Company name': 'Windesco',
        'Category': 'Solar PV',
        'Status': 'Met but no live discussions',
        'Country': 'United Kingdom',
        'Description': 'Windesco makes blade inspection drones and AI-based damage detection software for onshore and offshore wind turbines, cutting inspection time from 4 hours to 25 minutes per turbine.',
        'WBWSite': 'https://windesco.com',
        'Added by': 'Sarah Mitchell',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Key contact': 'Lena Fischer',
        'Key contact email': 'l.fischer@windesco.com',
        'Relationship owner': 'Sarah Mitchell',
        'Avoid': 'No',
        'BUs': ['R&D', 'EDF Power Solutions'],
        'Notes': '2024-05-07: Good product. Renewables team keen. No active programme to deploy into right now.',
    },

    # ── No contact ─────────────────────────────────────────────────────────
    {
        'Company name': 'Dalkia Biomass Tech',
        'Category': 'Sustainable Fuels & Biomass',
        'Status': 'No contact',
        'Country': 'France',
        'Description': 'Develops compact biomass gasification units for combined heat and power in industrial parks, targeting 500kW–2MW sites with a 7-year payback.',
        'WBWSite': 'https://dalki-biomass.fr',
        'Added by': 'AI (similar to Biocraft Energy)',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Avoid': 'No',
        'BUs': ['Dalkia'],
    },
    {
        'Company name': 'Flexitricity',
        'Category': 'Smart Grid & Flexibility',
        'Status': 'No contact',
        'Country': 'United Kingdom',
        'Description': 'Flexitricity aggregates demand-side response from industrial and commercial sites — cold stores, data centres, manufacturers — to sell frequency response and reserve services to National Grid.',
        'WBWSite': 'https://flexitricity.com',
        'Added by': 'AI (similar to Gridshift)',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Avoid': 'No',
        'BUs': ['EDF Power Solutions'],
    },
    {
        'Company name': 'Sunamp',
        'Category': 'Heat Pumps (Residential)',
        'Status': 'No contact',
        'Country': 'United Kingdom',
        'Description': 'Sunamp makes compact phase-change thermal storage units that integrate with heat pumps and solar to store heat as a solid-to-liquid transition, offering 5× the density of a hot water cylinder.',
        'WBWSite': 'https://sunamp.com',
        'Added by': 'AI (similar to PassivSystems)',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Avoid': 'No',
        'BUs': ['EDF Heat Pumps'],
    },
    {
        'Company name': 'Eatron Technologies',
        'Category': 'Battery Storage & BESS',
        'Status': 'No contact',
        'Country': 'United Kingdom',
        'Description': 'Eatron develops cloud-connected battery management systems (BMS) for second-life and stationary storage, extending pack life by 30% through adaptive cell balancing algorithms.',
        'WBWSite': 'https://eatron.com',
        'Added by': 'David Ferguson',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Avoid': 'No',
        'BUs': ['EDF Power Solutions', 'R&D'],
    },
    {
        'Company name': 'Orbital Marine Power',
        'Category': 'Marine & Tidal Energy',
        'Status': 'No contact',
        'Country': 'United Kingdom',
        'Description': 'Orbital Marine Power operates the world\'s most powerful tidal turbine, the O2, generating 2MW from tidal stream in the Orkney Islands and targeting commercial arrays from 2026.',
        'WBWSite': 'https://orbitalmarine.com',
        'Added by': 'Tom Whitfield',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Avoid': 'No',
        'BUs': ['EDF Power Solutions'],
    },
    {
        'Company name': 'Origami Energy',
        'Category': 'Smart Grid & Flexibility',
        'Status': 'No contact',
        'Country': 'United Kingdom',
        'Description': 'Origami Energy\'s platform connects distributed energy assets — batteries, EV chargers, generation — into a single optimised portfolio, trading across multiple markets simultaneously.',
        'WBWSite': 'https://origamienergy.com',
        'Added by': 'Rachel Osei',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Avoid': 'No',
        'BUs': ['EDF Power Solutions', 'Retail'],
    },

    # ── Avoid ──────────────────────────────────────────────────────────────
    {
        'Company name': 'Greenwash Corp',
        'Category': 'ESG & Sustainability Reporting',
        'Status': 'Met but no live discussions',
        'Country': 'United Kingdom',
        'Description': 'Claims to provide automated scope 3 emissions reporting but audit found significant methodology errors. Legal dispute with a previous client is ongoing.',
        'WBWSite': 'https://greenwash-corp.com',
        'Added by': 'David Ferguson',
        'EDF PV  portfolio ': 'No',
        'NDA signed': 'No',
        'Avoid': 'Yes',
        'BUs': [],
        'Notes': '2024-03-12: Do not engage. Methodology review found non-compliant reporting. Legal team aware.',
    },
]


def make_demo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Startups'

    # ── Header row styling ────────────────────────────────────────────────────
    header_fill   = PatternFill('solid', fgColor='10367A')   # EDF navy
    header_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border   = Border(
        bottom=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='EEEEEE'),
    )

    # Write headers (BOM on first col to match real file)
    for ci, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.value = ('﻿' if ci == 1 else '') + h
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column index lookup
    hmap = {h.lstrip('﻿').strip(): i + 1 for i, h in enumerate(HEADERS)}

    def w(row, header, value):
        col = hmap.get(header.strip())
        if col and value is not None:
            ws.cell(row=row, column=col).value = value

    # Status row colours
    STATUS_FILLS = {
        'Live collaboration':          'FFD9C2',
        'Advanced discussions':        'DDF4D2',
        'Early Exploration':           'DCEAFF',
        'Met but no live discussions': 'E0E0E0',
        'No contact':                  'F5F5F5',
    }

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, s in enumerate(STARTUPS, start=2):
        status = s.get('Status', 'No contact')
        row_fill = PatternFill('solid', fgColor=STATUS_FILLS.get(status, 'FFFFFF'))

        w(ri, 'Company name',      s['Company name'])
        w(ri, 'Category',          s.get('Category', ''))
        w(ri, 'Status',            status)
        w(ri, 'Country',           s.get('Country', ''))
        w(ri, 'Description',       s.get('Description', ''))
        w(ri, 'WBWSite',           s.get('WBWSite', ''))
        w(ri, 'Added by',          s.get('Added by', ''))
        w(ri, 'EDF PV  portfolio ', s.get('EDF PV  portfolio ', 'No'))
        w(ri, 'NDA signed',        s.get('NDA signed', 'No'))
        w(ri, 'Key contact',       s.get('Key contact', ''))
        w(ri, 'Key contact email', s.get('Key contact email', ''))
        w(ri, 'Relationship owner', s.get('Relationship owner', ''))
        w(ri, 'Avoid',             s.get('Avoid', 'No'))
        w(ri, 'Notes',             s.get('Notes', ''))
        w(ri, 'Notes 2',           s.get('Notes 2', ''))
        w(ri, 'Notes 3',           s.get('Notes 3', ''))

        # Business unit checkboxes
        for bu in s.get('BUs', []):
            col = hmap.get(bu)
            if col:
                ws.cell(row=ri, column=col).value = True

        # Light row tint
        for ci in range(1, len(HEADERS) + 1):
            ws.cell(row=ri, column=ci).border = Border(
                bottom=Side(style='thin', color='F0F0F0'),
                right=Side(style='thin', color='F0F0F0'),
            )

    # ── Column widths ─────────────────────────────────────────────────────────
    WIDTH_MAP = {
        'Company name': 28,
        'Description': 60,
        'Notes': 45, 'Notes 2': 45, 'Notes 3': 45,
        'Category': 32,
        'WBWSite': 30,
        'Status': 26,
        'Country': 16,
        'Added by': 18,
        'Key contact': 20,
        'Key contact email': 28,
        'Relationship owner': 20,
    }
    for header, width in WIDTH_MAP.items():
        col = hmap.get(header)
        if col:
            ws.column_dimensions[get_column_letter(col)].width = width

    # BU columns narrow
    for bu in BU_COLS:
        col = hmap.get(bu)
        if col:
            ws.column_dimensions[get_column_letter(col)].width = 5

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = 'A2'

    wb.save(OUT)
    print(f'✓ Saved {len(STARTUPS)} demo startups to: {OUT}')


if __name__ == '__main__':
    make_demo()
