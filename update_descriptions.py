import openpyxl
import random

wb = openpyxl.load_workbook("Startup database.xlsx")
ws = wb.active

replacements = {
    7: "Aerleum develops bifunctional materials for combined atmospheric CO2 capture and conversion to methanol in a single reactor. Its proprietary adsorbents capture CO2 at ambient temperature and low concentrations, while optimised catalysts convert it to methanol. The integrated reactor eliminates energy-intensive compression, transport, and storage steps, enabling cost-competitive carbon-neutral fuel production at industrial scale.",

    9: "Allegiant H2O connects investors with water infrastructure projects globally, bridging the gap between project financing and the technologies needed to deliver them. It operates as a platform to accelerate development of water-related projects in regions facing scarcity and inadequate infrastructure.",

    10: "Allye Energy provides intelligent battery storage for on-grid and off-grid applications, targeting construction, film, utilities, and EV charging sectors. Its systems are claimed to be 2x cheaper than alternatives in Europe and the US, with grid-forming capabilities. The company is developing Energy Banking, a cloud-hosted shared battery infrastructure model aimed at reducing energy storage costs at scale.",

    15: "Aquaprox provides services across the full water cycle, including auditing and consulting on water installations, production and formulation of water treatment chemicals, engineering and assembly services.",

    17: "Aquasys offers a SaaS Water Resource Management System for climate-induced water risk. It includes modules for data collection, AI-driven analytics, and application-specific data use such as irrigation and drinking water management. The platform is designed as a turnkey solution for resource managers without extensive IT infrastructure.",

    20: "Ask for the Moon is a Q&A platform that surfaces tacit knowledge within enterprises using machine learning. It integrates into existing workflows, adapts to each business case, and aims to reduce time spent searching for internal knowledge by connecting employees with the right colleagues.",

    24: "BeZero Carbon is a carbon ratings agency providing independent, risk-based, project-level ratings for carbon credits. Its ratings and risk tools are used by organisations to assess carbon project quality and support investment decisions in environmental markets.",

    26: "BioEsol provides an intelligent battery energy storage system with accompanying software to monitor energy generation and manage stored power. It targets businesses facing energy instability, high costs, or voltage variations, and integrates renewable energy sources to deliver uninterrupted power supply.",

    30: "Blue Water Intelligence (BWI) is a Toulouse-based company specialising in monitoring inland surface water reserves. It combines in-situ data, satellite observations, and machine learning to power hydrological models and deliver forecasts on water availability, quality, floods, and droughts to help organisations manage climate-related water risk.",

    31: "Boreales Energy converts electrical energy, including from renewable sources, into thermal energy stored as ice. Its high-efficiency submerged evaporator produces and stores ice for deferred, low-cost cold production, enabling flexible cooling consumption timed to suit operational and economic needs.",

    37: "Carbon Re develops AI software to decarbonise energy-intensive industries such as cement, steel, and glass. Its Delta Zero platform uses deep learning to enable rapid development and deployment of low-carbon industrial processes, targeting sectors that collectively account for over 20% of global emissions.",

    38: "Carbon Upcycling permanently stores CO2 in industrial byproducts and minerals, transforming them into high-performance supplementary materials for cement and concrete. Its patented technology targets hard-to-abate industries, diverts industrial waste from landfill, and supports circular economy goals.",

    42: "Chainels is a tenant experience platform for multi-tenant commercial, residential, and transport destinations. Its app enables property managers and tenants to handle communication, amenities, data reporting, and marketing. The company serves over 400 communities across 15 countries, including clients such as CBRE, Wereldhave, and Amvest.",

    44: "Chemdoc Water Technologies manufactures high-technology equipment for industrial water filtration, purification, and demineralisation, as well as drinking water treatment and seawater desalination. It also provides expertise in industrial cooling systems and hot water network disinfection, serving industrial clients, municipalities, and building managers.",

    45: "Chiral Energies develops quantum-chemistry-based nano-coatings that enhance the performance of electrochemical devices including electrolysers, batteries, and fuel cells. The coatings are applied via a low-cost process compatible with existing technologies, improving capacity, efficiency, and lifespan.",

    46: "Circu Li-ion is a Luxembourg-based battery upcycling company. Its automated process uses AI and machine vision to recover up to 90% of end-of-life battery cells and components, avoiding premature recycling of functional parts. The approach reduces battery-related CO2 emissions by 48% compared to a traditional battery lifecycle, serving EV, micromobility, and power tools industries.",

    53: "CTRL Energy develops energy storage and battery management systems targeting the transition to full electrification. It builds advanced energy storage and supply systems with no comparable market capacity or autonomy, using proprietary software and hardware to deliver aggressive pricing.",

    65: "Enerbrain provides a smart retrofit system for non-residential buildings, combining IoT hardware and cloud algorithms to optimise energy use in real time and deliver around 30% operational savings. The platform also monitors indoor air quality and provides occupant engagement tools.",

    68: "EnjoyElec is an energy management app that optimises electricity usage across home devices including batteries, solar panels, and the grid. It prioritises solar and battery use during high generation, exploits negative grid prices, and uses Edge AI via its controller to personalise energy strategies.",

    69: "Enode provides APIs that connect users' energy devices — including EVs, solar inverters, home batteries, and thermostats — to third-party applications. The platform reduces integration complexity for developers building energy management apps, covering device connectivity and ongoing maintenance.",

    70: "Enspired is a digital energy trading-as-a-service company that uses AI algorithms to trade flexible energy assets on power spot markets across Europe. Its fully automated trading system optimises and monetises client flexibility from renewables, batteries, and consumption assets while supporting grid stability.",

    75: "Epoch Biodesign uses AI-engineered enzymes to recycle nylon 6,6 and other plastics with no sustainable end-of-life pathway today. Backed by $50M from Lowercarbon Capital, Extantia, and Inditex, it holds commercial partnerships with major apparel brands and operates a lab and pilot plant in London processing 200 tonnes of waste per year.",

    76: "Evergen develops software to optimise solar and battery system performance for homeowners and orchestrate large fleets of batteries into Virtual Power Plants (VPPs). Its platform enables decentralised energy networks where distributed solar and battery systems supply excess power to the grid, currently deployed in Australia.",

    79: "Finmile is an AI-powered delivery execution platform that manages the full delivery day end-to-end. It dynamically optimises routes in real time, predicts failures, enforces proof of delivery, and automates billing. Used by enterprise retailers and delivery networks as a single operating system replacing fragmented logistics tools.",

    80: "First Light Fusion is an Oxford University spin-out researching inertial confinement fusion for energy generation. It combines power generation theory, numerical simulation, and experimental validation, collaborating with institutions including Oxford, Warwick, UCL, and Imperial College London.",

    81: "Fornax offers a lease-based financing product for heat pump installations, covering equipment, installation, and full maintenance from around £60/month for up to 12 years. Its proprietary underwriting model automates engineering design, optimises capital spend, and prices installation and maintenance risk across a broad range of property types.",

    84: "From Sand to Green creates solar-powered desalination-fed plantations that restore soil fertility using agroforestry techniques. Structured as SPVs, these plantations produce food, biofuel, and carbon credits by aggregating innovations including containerised solar desalination, biochar, drip irrigation, and satellite analysis.",

    85: "Fusebox Energy builds virtual power plant layers globally, trading demand flexibility to energy markets and balancing services. It provides technology and expertise to local companies with controllable flexibility or access to flexible client portfolios, operating as a marketplace for flexibility and related services.",

    93: "Green Energy Options designs and produces in-home energy displays, online energy services, and mobile applications for smart home energy management. Since 2006 it has developed over 25 products and sold more than 4 million energy monitoring systems, serving utility companies and retail partners globally.",

    95: "H2CHP is developing a modular free-piston linear generator delivering fuel-cell-level electrical efficiency at engine-generator cost, with multi-fuel capability and a compact footprint. The platform targets resilient, low-carbon distributed power for ports, maritime infrastructure, and off-grid industrial sites.",

    100: "HETWA is a water and soil consultancy providing diagnostics and tailored solutions for water management challenges. Its innovation division develops the Mewiso web platform, which enables real-time monitoring and GIS-based visualisation of water resource quality and quantity across multiple management scales including catchments and river basin districts.",

    103: "Hydrogrid offers a SaaS platform that helps renewable energy plant owners optimise dispatch and electricity market participation. It combines hydrological forecasting using machine learning with automated energy marketing tools, reducing daily operational workload for power plant operators.",

    104: "Hyperion Robotics provides automation and additive manufacturing solutions for infrastructure and commercial construction projects. Its systems produce durable reinforced concrete structures for a range of applications. The company was founded in 2019 and is based in Espoo, Finland.",

    106: "Infyos provides a digital platform for battery and EV manufacturers to manage and improve supply chain sustainability. It helps companies meet customer and regulatory deadlines by combining deep domain knowledge in battery materials with software that identifies actionable sustainability improvements.",

    111: "Juicy is an energy tech venture building on the founding team's experience from elmo (electric car subscription, sold to Constellation Automotive) and heata/Deep Green (distributed heat-reuse compute, acquired by Octopus Energy Generation). The team of four brings expertise in scaling energy businesses from zero to exit.",

    112: "Kapacity.io provides an electricity load balancing service for buildings, shifting consumption to periods of lower prices and higher renewable output. By aggregating buildings into a Virtual Power Plant, it participates in grid balancing and helps reduce both electricity costs and CO2 emissions.",

    113: "Kavaken is a UK-based climate tech company providing Operational AI for renewable energy assets. Its platform connects to SCADA, condition monitoring, and financial systems to detect underperformance, predict failures, and quantify revenue impact. It currently supports multiple gigawatts of wind capacity across Europe.",

    116: "Kumulus Water produces a compact atmospheric water generator (1m x 3m, 30 kg) equipped with a battery, filter, and monitoring system. It condenses and filters water from air, can be paired with solar power, and delivers water at 30% lower cost than bottled water. Business model covers unit sales and maintenance, targeting NGOs.",

    119: "Laiout is a multi-player design tool for the architecture and construction industry that generates floor plans in milliseconds while adhering to architectural and regulatory constraints. Backed by Proptech VCs, its platform optimises the early design phase, saving up to 20 weeks compared to traditional processes.",

    127: "Mimbus ARXR provides AR/VR solutions for industrial training and applications.",

    130: "Modern Synthesis is a London-based material innovation company developing a PU-free, cellulose-based coating technology to replace polyurethane and PVC-based synthetic leathers in fashion. Its roll-to-roll compatible process integrates at the Tier 2 manufacturing level without infrastructure overhauls, enabling supply chain adoption. The company has raised €9.7M from investors including Extantia and Artesian VC.",

    131: "MX3D develops software for metal additive manufacturing, with applications in the nuclear sector including work with Framatome.",

    133: "Natural Building Systems (NBS) uses bio-based crops and modern manufacturing to produce its ADEPT modular construction system, designed for fast, affordable, and circular housing. Digital traceability ensures components can be reused or repurposed. NBS is developing regional partnerships and pilot projects across the UK.",

    134: "NatureMetrics provides a biodiversity monitoring platform using environmental DNA to identify and quantify organisms at large spatial and temporal scales. Its platform combines biodiversity science and molecular methods for applications ranging from conservation to environmental impact assessment.",

    135: "NetZeroNitrogen is developing a biological alternative to synthetic nitrogen fertiliser, targeting the $200bn fertiliser market and over 1 billion tonnes of CO2e in annual emissions. Field trials are underway for a first product for rice, expected to launch by end of 2026, following two funding rounds totalling $8.5M.",

    136: "Nido streamlines heat pump installation in the residential sector, offering clean heating and cooling at affordable pricing. Its platform removes time-consuming tasks for installers and helps homeowners transition away from fossil fuel boilers, targeting the 65M+ over-life boilers in the EU.",

    140: "Oklima develops and finances carbon capture and low-carbon projects using the carbon credit system. It quantifies and certifies carbon reductions across nature-based, agricultural, and technological project types in France, Europe, and globally, and develops new methodologies for novel project categories.",

    142: "Openvolt is building an API for smart meter data, integrating with smart meter data sources across Europe to provide developer-friendly access for application builders. Founded by Dave Curran and Don O'Leary, it addresses regulatory and technical barriers to accessing the data of over 1 billion deployed smart meters globally.",

    147: "Origen Carbon develops a zero-emission lime kiln platform using a patented oxy-fuel flash calciner demonstrated at over 3,500 tpa. The technology produces highly reactive lime without requiring electrification, enabling decarbonisation across cement, steel, drinking water, and carbon removal pathways. Origen has a customer pipeline including Fortera, MLC, Inpex, and Shell/Mitsubishi Corp.",

    149: "Oxylus Energy converts CO2 to methanol via electrolysis at room temperature and pressure using a breakthrough electrochemical catalyst. Its modular e-methanol technology offers high energy efficiency and emission reduction potential, enabling on-site decarbonisation for industrial applications.",

    153: "Podero is a SaaS platform enabling utilities to offer behind-the-meter services and manage demand flexibility. It integrates with major steerable devices (heat pumps, EVs, batteries, inverters) across 51 brands and supports trading on electricity markets. Current customers include E.ON and KELAG, with €30.5k MRR.",

    155: "Porpoise Power is an Oxford University spin-out developing tidal or wave energy technology. It raised a £1.2M pre-seed round led by Zero Carbon Capital in September 2024 and is progressing through prototype testing stages, with a second prototype under development alongside a seed fundraise in 2025.",

    157: "PowerUp is a CEA spin-off that develops battery optimisation and health monitoring technology for lithium-ion batteries. Its solution extends battery lifespan and enables remote state-of-health monitoring, supporting fleet management for electric transport. Current clients include the French DGA and Legrand.",

    160: "Pyri is developing a patent-pending, non-toxic organic sensor for ultra-early wildfire detection that requires no battery, maintenance, or connectivity at the sensing point. The air-deployable system is designed for large-scale perimeter coverage at lower cost than satellite, camera, or drone alternatives, targeting high-risk regions globally.",

    163: "Reality Agency is a virtual reality training agency offering consulting, solution development, and platform and catalogue services for industrial VR training.",

    165: "Reefilla provides a mobile, predictive EV charging delivery service, bringing charging to vehicles wherever they are parked. It uses 100% green energy, fully electric delivery vans, and second-life batteries, operating on a circular model to support net-zero emissions.",

    167: "Regli Energy Systems manufactures high-efficiency heat pumps for domestic and commercial use in Switzerland and Germany, with heating output ranging from 4 kW to 1,100 kW. The company focuses on maximising heat output per unit of electricity, targeting older, poorly insulated buildings transitioning away from fossil fuel heating.",

    168: "Rensair has created an air quality ecosystem that retrofits existing HVAC systems to reduce carbon emissions and energy use by over 40% while improving indoor air quality. Its certified technology targets large corporations seeking to meet ESG and Net Zero commitments by cutting the carbon footprint of building ventilation.",

    170: "Repath provides a SaaS platform for climate risk assessment and adaptation planning. It delivers location-specific forecasts of future physical climate risks, identifies asset vulnerabilities, and recommends adaptation measures, with results accessible via the platform or API.",

    175: "Shayp is a Belgian IoT water monitoring company specialising in consumption analysis and early leak detection. Its non-invasive technology, labelled by the Solar Impulse Foundation, helps property owners and managers achieve an average 20% reduction in water bills and prevent water damage.",

    179: "SkyVisor provides drone-based industrial inspection for renewable energy infrastructure, including wind farms and solar installations. It tracks wear and degradation over time to support preventive maintenance planning, delivering detailed structural health assessments of energy assets.",

    180: "SmartEnds provides IoT-based smart waste management under its BrighterBins product, combining modular sensors (camera, laser, ultrasonic, GPS) with cloud services and edge computing. It offers routing, prediction, and image analysis capabilities via API or its own software applications.",

    182: "SmartShift Energy is developing a consumer-owned home energy management solution that monitors and optimises multiple devices to reduce electricity costs and provide demand-side flexibility to the grid, targeting the residential sector.",

    183: "SMPnet develops grid optimisation and management software using digital twin and physics-based models, multi-vendor data translation, and real-time dynamic control. Its platform manages renewable integration, congestion, and regulatory compliance, sold on a perpetual licence-per-substation model with support contracts.",

    185: "SoHHytec produces a photo-electrochemical device that uses sunlight and water to generate hydrogen, oxygen, electricity, and heat simultaneously on-site. Its integrated system offers high solar-to-fuel and solar-to-electricity efficiency with minimal space requirements, targeting decentralised clean hydrogen production.",

    186: "SolCold has developed a patented nano-technological coating material that cools surfaces using sunlight, without electricity. Applied to cars, buildings, containers, and aircraft, the material converts heat into radiation via anti-Stokes fluorescence, providing a passive cooling effect that intensifies as sunlight increases.",

    187: "Soldera provides an AI-powered platform that automates the sales and administration of Guarantees of Origin (GOs) and REGOs for renewable energy producers, reducing administrative work by up to 95% and improving revenue. Since launch it has managed 372 GWh across Estonia, Latvia, and Lithuania, with 178 customers and €240K GMV.",

    189: "Soly is a European clean energy company offering solar panels, batteries, and EV charging stations to residential and business customers. Founded in 2013, it operates in the Netherlands, South Africa, Belgium, Germany, and the UK with 140 employees. It is B Corp certified and a Deloitte Technology Fast50 company.",

    190: "Sorted Technologies builds autonomous waste-sorting and measurement systems for the recycling industry, combining computer vision, robotics, and data analytics. Its Sortbot robot (TRL-7), Sorted Lights sensors, and Waste Audit Studio analytics platform are deployed commercially with clients including SUEZ and Bywaters.",

    193: "Spotr is an automated building inspection platform that aggregates data from multiple sources to produce a complete description of buildings, including roof size and orientation, windows, and walls, using satellite imagery. EDF Pulse Ventures is an investor.",

    197: "Strong by Form develops Woodflow, a design-to-fabrication technology combining AI-driven structural optimisation with proprietary additive manufacturing to produce high-performance lightweight biocomposites from timber. The material is positioned as an alternative to aluminium, fibre-reinforced plastics, steel, and concrete for vehicles and buildings.",

    200: "Sweetch Energy has developed a nanofluidic membrane for osmotic energy generation that achieves power output around 1 kW per square metre of membrane — approximately 1,000 times higher than conventional salinity-gradient technologies. The membrane enables selective ion transport between freshwater and saltwater flows, converting the salinity gradient into electrical current.",

    201: "SWITCH provides an AI-powered urban mobility platform integrating its PULSE-AI generative engine and OPTIMESH routing algorithm. The platform generates and forecasts urban mobility scenarios, supports fleet management, optimises parking and EV charging station placement, and offers real-time compliance monitoring for mobility providers, logistics firms, and local governments.",

    203: "Tenko was founded in 2021 to quantify the impacts of climate volatility using machine learning and big data. Its team combines experience in data analysis and quantitative risk modelling, with a focus on climate change risk analytics.",

    204: "TerraWaste is a chemical recycling company using its htloop hydrothermal liquefaction technology to convert non-recyclable plastic waste into crude oil, biochar, and high-value chemicals. The continuous-flow system reuses heat and water internally, producing carbon-negative outputs. TerraWaste is the first company to commercialise HTL for plastic-to-plastic recycling.",

    208: "Tipi Energy is an EDF spin-out offering financial instruments backed by electricity, structured as mini-PPAs for residential and business customers. Products provide upfront payment, long-term protection against electricity price increases, supplier flexibility, and contract resale options, functioning as an energy price hedging mechanism.",

    209: "Tree Water provides wastewater treatment and recycling services, specialising in advanced oxidation, micropollutant treatment, and chemical oxygen demand reduction. Active projects include laundry wastewater recycling (Life Recyclo), magnetite-based micropollutant treatment (Magnet'eau), energy-efficient oxidation (Electrotate), and snow production from treated wastewater, several developed with CNRS.",

    211: "True Elements operates a water quality forecasting platform that aggregates data from public water sensors to help businesses, governments, and financial institutions assess localised water quality, quantity, and climate risk. The platform normalises and processes watershed data to produce water quality and climate risk forecasts.",

    212: "Turbit develops and deploys AI monitoring for wind parks, enabling operators and asset managers to detect abnormal turbine behaviour early. The platform identifies power anomalies, predicts major component failures months ahead, monitors repair quality, and improves detection accuracy through user feedback.",

    215: "Urbio is a software platform that helps actors in the energy transition plan and design energy infrastructure for building heating, cooling, and electrification. It combines human expertise and AI to replace manual trial-and-error in energy system design, improving outcomes and significantly reducing design time.",

    217: "Vasco offers an equity-based financing solution for whole-home energy renovation in France. Vasco funds 100% of renovation costs in exchange for a share of the property. The homeowner remains in their home, pays no interest or instalments, and repays Vasco upon sale or earlier if preferred.",

    218: "Vertus Energy is an industrial biotechnology company that uses electrical stimulation to control bacterial behaviour in anaerobic digestion, increasing biomethane yields by 30–60% from the same feedstock. Its modular Biotransformation Platform integrates into existing or new commercial plants with minimal operational disruption.",

    219: "VivaDrive helps companies transition their vehicle fleets to electric vehicles. Its platform analyses transport tasks, vehicles, and drivers to model EV deployment scenarios and build business cases for charging infrastructure. It also manages mixed electric and traditional fleets and charging infrastructure post-transition.",

    220: "Vuala provides automatic food waste separation technology combining mechanical design with micro-organisms and bacteria to convert food waste into energy feedstock on-site within hours. The system reduces odour, allows months of storage, and delivers up to 95% savings in labour and logistics costs, with deployments at Heathrow Airport and Amazon.",

    221: "WAS Company develops sustainable building materials with negative carbon footprint using transformed plastic, construction, and mining waste. It serves the mining and cement industries by providing low-emission materials with competitive performance, operating as a science-based company for circular construction solutions.",

    222: "WASE converts industrial wastewater and sludge into biogas on-site using its electro-methanogenic reactor (EMR), the world's largest commercial unit of its type. It targets food and beverage manufacturers, helping them reduce operating costs and meet environmental discharge standards. Partnerships include Engie and Kimmeridge, with over 1 TWh of projects in pipeline.",

    224: "WaterShed Monitoring helps organisations that abstract, distribute, treat, and discharge water to optimise their operations.",

    225: "WATURA is developing an e-learning platform for technical water industry training, delivering short video modules, quizzes, and exercises on topics including electrical systems, fluid mechanics, and chemistry. The personalised subscription-based platform targets public water and sanitation services in France and internationally, with planned expansion into other technical sectors.",

    226: "Wayout International produces water microfactories that purify any local water source using double reverse osmosis, desalination, UV treatment, and remineralisation. Solar-powered and cloud-connected, each unit supplies drinking water for around 300 households daily while eliminating up to 250,000 plastic bottles and 25 tonnes of CO2 per month.",

    230: "XSun designs solar-powered autonomous aerial vehicles capable of continuous day-and-night flight, replacing manned aircraft and helicopters for surveillance missions. Its aerodynamic solution integrates GaAs photovoltaic cells in composite materials with embedded avionics, energy, and communications systems.",

    232: "Yokahu is a Lloyd's of London coverholder and automated insurance intermediary selling parametric insurance policies triggered by proven events rather than assessed losses. Its first product is hurricane insurance for Caribbean islands, available in 90+ countries, with technology automating the full insurance experience.",
}

for row, new_desc in replacements.items():
    ws.cell(row=row, column=6).value = new_desc

wb.save("Startup database.xlsx")
print("Saved successfully.")

# Verify 5 random rows
sample_rows = random.sample(list(replacements.keys()), 5)
print("\n--- Verification (5 random rows) ---")
wb2 = openpyxl.load_workbook("Startup database.xlsx")
ws2 = wb2.active
for row in sorted(sample_rows):
    company = ws2.cell(row=row, column=1).value
    desc = ws2.cell(row=row, column=6).value
    print(f"\nRow {row}: {company} ({len(desc)} chars)")
    print(desc)
