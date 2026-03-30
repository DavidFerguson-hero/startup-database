from flask import Flask, render_template, jsonify, request, session, redirect, url_for, send_file
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
import json, os, re, subprocess, secrets, math, threading
import openpyxl
from datetime import datetime
from io import BytesIO
import ai_tasks

app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True
BASE = os.path.dirname(os.path.abspath(__file__))
# DATA_DIR: override with env var to point at an EFS mount or persistent volume on AWS
DATA_DIR = os.environ.get('DATA_DIR', BASE)
EXCEL = os.path.join(DATA_DIR, 'Startup database.xlsx')
UPDATES = os.path.join(DATA_DIR, 'updates.json')   # legacy — migrated on first run
STARTUPS_DIR = os.path.join(DATA_DIR, 'startups')
MIGRATED_FLAG = os.path.join(DATA_DIR, '.notes_migrated')
PW_FILE = os.path.join(DATA_DIR, '.password_hash')
SK_FILE = os.path.join(DATA_DIR, '.secret_key')
COLLECTIONS_FILE = os.path.join(DATA_DIR, 'collections.json')

# ── Auth setup ───────────────────────────────────────────────────────────────

# Secret key — env var takes priority (for Railway/cloud), else persist locally
if os.environ.get('SECRET_KEY'):
    app.secret_key = os.environ['SECRET_KEY']
elif os.path.exists(SK_FILE):
    app.secret_key = open(SK_FILE).read().strip()
else:
    app.secret_key = secrets.token_hex(32)
    try:
        open(SK_FILE, 'w').write(app.secret_key)
    except OSError:
        pass  # ephemeral FS (AWS container) — set SECRET_KEY env var for stable sessions

# Password — env var APP_PASSWORD (plain text) takes priority over hash file
DEFAULT_PASSWORD = 'edf2024'
def check_password(pw):
    env_pw = os.environ.get('APP_PASSWORD')
    if env_pw:
        return pw == env_pw
    if os.path.exists(PW_FILE):
        return check_password_hash(open(PW_FILE).read().strip(), pw)
    return pw == DEFAULT_PASSWORD

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.path.startswith('/api/'):
                return jsonify({'ok': False, 'error': 'session_expired'}), 401
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated

STATUS_OPTIONS = ['No contact', 'Met but no live discussions', 'Early Exploration', 'Advanced discussions', 'Live collaboration']
NOTE_PREFIX = 'Notes'  # column headers: "Notes", "Notes 2", "Notes 3", …

# Checkbox columns B–L that represent business unit membership
BU_CHECKBOX_COLS = [
    'Nuclear operations', 'Nuclear Services', 'Dalkia', 'BWS', 'Retail',
    'EDF Heat Pumps', 'Contact Solar', 'Hynamics', 'R&D', 'EDF Power Solutions', 'Pod',
]

def is_checked(val):
    """Excel checkbox cells can come through as bool True, string 'True', or numeric 1."""
    if val is True: return True
    if isinstance(val, str) and val.strip().lower() == 'true': return True
    try:
        return float(val) == 1.0
    except (TypeError, ValueError):
        return False

def folder_name(name):
    return re.sub(r'[/\\:*?"<>|]', '-', str(name)).strip()

def startup_folder(name):
    return os.path.join(STARTUPS_DIR, folder_name(name))

# Sub-category → broad theme mapping
SUBCAT_THEME = {
    'Water Technology & Management':        'Environment & Sustainability',
    'Biodiversity & Nature Restoration':    'Environment & Sustainability',
    'Circular Economy & Waste':             'Environment & Sustainability',
    'Carbon Capture & Storage':             'Environment & Sustainability',
    'Climate Risk & Resilience':            'Environment & Sustainability',
    'Agri & FoodTech':                      'Environment & Sustainability',
    'ESG & Sustainability Reporting':       'Environment & Sustainability',
    'Carbon Offsetting & Credits':          'Environment & Sustainability',
    'Heat Pumps (Residential)':             'Buildings & Cities',
    'Building Energy Management':           'Buildings & Cities',
    'District Heating & Cooling':           'Buildings & Cities',
    'Digital Construction & Real Estate':   'Buildings & Cities',
    'Home Energy Management & Smart Metering': 'Buildings & Cities',
    'EV Charging & Fleet Electrification':  'Buildings & Cities',
    'Retrofit & Green Building Finance':    'Buildings & Cities',
    'Heat Pumps (Industrial & Commercial)': 'Buildings & Cities',
    'Demand Response & Flexibility':        'Energy Storage & Flexibility',
    'Smart Grid & Grid Management':         'Energy Storage & Flexibility',
    'Battery Storage & BESS':               'Energy Storage & Flexibility',
    'Long-Duration Energy Storage':         'Energy Storage & Flexibility',
    'Virtual Power Plant & Aggregation':    'Energy Storage & Flexibility',
    'Circular Battery Economy':             'Energy Storage & Flexibility',
    'AI & Data Analytics':                  'AI & Digital',
    'Inspection, Drones & Robotics':        'Industry & Manufacturing',
    'Industrial Decarbonisation':           'Industry & Manufacturing',
    'Solar PV':                             'Renewables',
    'Wind Energy':                          'Renewables',
    'Marine & Tidal Energy':               'Renewables',
    'Space-Based Solar Power':              'Renewables',
    'Floating Wind':                        'Renewables',
    'Hydrogen Production':                  'Hydrogen & Fuels',
    'Sustainable Fuels & Biomass':          'Hydrogen & Fuels',
    'Nuclear Operations & Digital Tools':   'Nuclear & Deep Tech',
    'Nuclear Training & VR/AR':             'Nuclear & Deep Tech',
    'Nuclear Decommissioning & New Tech':   'Nuclear & Deep Tech',
    'Energy Trading & Optimisation':        'Energy (General)',
    'Energy Investment & Project Finance':  'Services',
    'HR, Training & Professional Services': 'Services',
}

def get_theme(category, description=None):
    if category:
        theme = SUBCAT_THEME.get(str(category).strip())
        if theme:
            return theme
    return 'Uncategorised'

# ── Note helpers ────────────────────────────────────────────────────────────

def note_header(n):
    """Column header for note number n (1-based): 'Notes', 'Notes 2', 'Notes 3'…"""
    return NOTE_PREFIX if n == 1 else f'{NOTE_PREFIX} {n}'

def is_note_header(h):
    h = str(h).strip()
    # Match "Notes", "Notes 2", "Notes 3" — also pandas duplicate suffixes "Notes.1", "Notes.2"
    if h == NOTE_PREFIX:
        return True
    if h.startswith(NOTE_PREFIX + ' ') and h[len(NOTE_PREFIX):].strip().isdigit():
        return True
    if h.startswith(NOTE_PREFIX + '.') and h[len(NOTE_PREFIX) + 1:].isdigit():
        return True
    return False

def parse_note_cell(value):
    """Parse note cell → {date, author, text}. Supports both legacy and authored formats."""
    if value is None:
        return None
    import math
    if isinstance(value, float) and math.isnan(value):
        return None
    s = str(value).strip()
    if not s or s.lower() == 'nan':
        return None
    # New format: "YYYY-MM-DD (Author Name): text"
    m = re.match(r'^(\d{4}-\d{2}-\d{2})\s+\(([^)]+)\):\s*(.*)', s, re.DOTALL)
    if m:
        return {'date': m.group(1), 'author': m.group(2), 'text': m.group(3).strip()}
    # Legacy format: "YYYY-MM-DD: text"
    if len(s) >= 12 and s[10] == ':':
        return {'date': s[:10], 'author': None, 'text': s[12:].strip()}
    return {'date': '', 'author': None, 'text': s}

def extract_website(raw):
    """Extract a clean URL from WBWSite cells that may contain multi-line crunchbase dumps."""
    if not raw:
        return None
    s = str(raw).strip()
    if not s or s == '-':
        return None
    # Multi-line format: "wBWSite => https://..." or "WBWSite => ..."
    for line in s.splitlines():
        if '=>' in line:
            key, _, url = line.partition('=>')
            if 'wbwsite' in key.lower() or 'website' in key.lower():
                url = url.strip()
                if url.startswith('http'):
                    return url
    # If no labelled line, return first http URL found
    for line in s.splitlines():
        if '=>' in line:
            _, _, url = line.partition('=>')
            url = url.strip()
            if url.startswith('http'):
                return url
    # Plain URL
    if s.startswith('http'):
        return s
    return None

def format_note_cell(date, text, author=None):
    if author:
        return f"{date} ({author}): {text}"
    return f"{date}: {text}"

# ── openpyxl helpers ────────────────────────────────────────────────────────

def find_excel_columns(ws):
    """Scan rows 1-4 for header. Returns (header_row, col_map) — col_map maps logical key → col index."""
    header_row = None
    col_map = {}
    for r in range(1, 5):
        for c in range(1, 60):
            val = ws.cell(row=r, column=c).value
            if not val:
                continue
            key = str(val).strip().lstrip('\ufeff').lower()
            if 'company' in key:
                col_map['name'] = c
                header_row = r
            elif key == 'status':
                col_map['status'] = c
        if header_row:
            break
    return header_row, col_map

def get_note_col_indices(ws, header_row):
    """Return sorted list of column indices for all Note* columns."""
    cols = []
    for c in range(1, 80):
        val = ws.cell(row=header_row, column=c).value
        if val and is_note_header(val):
            cols.append(c)
    return sorted(cols)

def ensure_note_col(ws, header_row, n):
    """Return column index for note #n, creating the header cell if needed."""
    header = note_header(n)
    for c in range(1, 80):
        val = ws.cell(row=header_row, column=c).value
        if val and str(val).strip() == header:
            return c
    # Create: place after the last used header column
    last = max((c for c in range(1, 80) if ws.cell(row=header_row, column=c).value is not None), default=1)
    new_col = last + 1
    ws.cell(row=header_row, column=new_col).value = header
    return new_col

def find_company_row(ws, header_row, name_col, name):
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        if v and str(v).strip() == name:
            return r
    return None

# ── Migration ───────────────────────────────────────────────────────────────

def migrate_notes_once():
    """On first run: write all updates.json notes into Note columns in the spreadsheet."""
    if os.path.exists(MIGRATED_FLAG):
        return
    legacy = {}
    try:
        with open(UPDATES) as f:
            legacy = json.load(f)
    except Exception:
        pass
    open(MIGRATED_FLAG, 'w').close()   # mark done even if nothing to migrate
    if not legacy:
        return
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb.active
        header_row, col_map = find_excel_columns(ws)
        if not header_row:
            return
        name_col = col_map['name']
        changed = False
        for company, notes in legacy.items():
            if not notes:
                continue
            row = find_company_row(ws, header_row, name_col, company)
            if row is None:
                continue
            # Check if any note cells already filled (don't overwrite)
            existing = get_note_col_indices(ws, header_row)
            if any(ws.cell(row=row, column=c).value for c in existing):
                continue
            # notes in updates.json are newest-first; write oldest-first into columns
            for i, note in enumerate(reversed(notes)):
                col = ensure_note_col(ws, header_row, i + 1)
                ws.cell(row=row, column=col).value = format_note_cell(note['date'], note['text'])
            changed = True
        if changed:
            wb.save(EXCEL)
    except Exception as e:
        print(f'[migration] Error: {e}')

# ── Data loading ────────────────────────────────────────────────────────────

def _clean(val):
    """Replace float NaN/Inf with None so jsonify never emits invalid JSON."""
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return None
    return val

def load_startups():
    df = pd.read_excel(EXCEL, header=0)
    df = df.where(pd.notna(df), None)
    df.columns = [c.lstrip('\ufeff').strip() for c in df.columns]

    # Identify note columns (sorted oldest→newest by column number)
    def note_sort_key(c):
        if c == NOTE_PREFIX:
            return 1
        suffix = c[len(NOTE_PREFIX):].strip().lstrip('.')
        try:
            return int(suffix)
        except ValueError:
            return 999

    note_col_names = sorted(
        [c for c in df.columns if is_note_header(c)],
        key=note_sort_key
    )

    result = []
    for _, row in df.iterrows():
        name = str(row['Company name']).strip() if row.get('Company name') else ''
        if not name or name.lower() in ('nan', 'none', 'company name'):
            continue

        # Sub-category is now the 'Category' column (col M)
        cat = str(row.get('Category') or '').strip() or None

        # Business units: read from Boolean checkbox columns B–L
        bus = [col for col in BU_CHECKBOX_COLS if is_checked(row.get(col))]
        business_unit = ', '.join(bus) if bus else None

        # EDF PV / NDA tag columns
        edf_pv = None
        nda = None
        for col in df.columns:
            if 'pv' in col.lower() and 'portfolio' in col.lower():
                edf_pv = row.get(col)
            if col.lower().strip() == 'nda signed':
                nda = row.get(col)

        # Collect notes (oldest→newest in columns, display newest-first)
        notes = []
        for col in note_col_names:
            parsed = parse_note_cell(row.get(col))
            if parsed:
                notes.append(parsed)
        notes.reverse()

        avoid_raw = str(row.get('Avoid') or '').strip()
        result.append({
            'name': name,
            'business_unit': business_unit,
            'category': cat,
            'tags': [cat] if cat else [],
            'primary_category': cat or 'Uncategorised',
            'theme': get_theme(cat),
            'country':           _clean(row.get('Country')),
            'description':       _clean(row.get('Description')),
            'status':            _clean(row.get('Status')),
            'status_comment':    _clean(row.get('Status comment')),
            'website': extract_website(row.get('WBWSite')),
            'added_by':          _clean(row.get('Added by')),
            'edf_pv': str(edf_pv).strip() if edf_pv else 'No',
            'nda':    str(nda).strip()    if nda    else 'No',
            'key_contact':       _clean(row.get('Key contact')),
            'key_contact_email': _clean(row.get('Key contact email')),
            'relationship_owner':_clean(row.get('Relationship owner')),
            'avoid': 'Yes' if avoid_raw.lower() == 'yes' else 'No',
            'updates': notes,
            'has_folder': os.path.isdir(startup_folder(name)),
        })
    return result

# ── Routes ──────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if check_password(request.form.get('password', '')):
            session['logged_in'] = True
            return redirect(request.args.get('next') or url_for('index'))
        error = 'Incorrect password'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/api/startups')
@login_required
def api_startups():
    return jsonify(load_startups())

@app.route('/api/status', methods=['POST'])
@login_required
def api_update_status():
    data = request.get_json()
    name = data.get('name', '').strip()
    new_status = data.get('status', '')
    if new_status not in STATUS_OPTIONS:
        return jsonify({'ok': False, 'error': 'Invalid status'})
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb.active
        header_row, col_map = find_excel_columns(ws)
        name_col = col_map.get('name')
        status_col = col_map.get('status')
        if not header_row or not name_col or not status_col:
            return jsonify({'ok': False, 'error': 'Could not locate columns'})
        row = find_company_row(ws, header_row, name_col, name)
        if row:
            ws.cell(row=row, column=status_col).value = new_status
            wb.save(EXCEL)
            return jsonify({'ok': True})
        return jsonify({'ok': False, 'error': 'Company not found'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/config')
@login_required
def api_config():
    return jsonify({
        'sharepoint_base': os.environ.get('SHAREPOINT_FOLDER_BASE', '').rstrip('/'),
    })

@app.route('/api/open-folder', methods=['POST'])
@login_required
def api_open_folder():
    """Fallback for local Mac use when no SharePoint base URL is configured."""
    data = request.get_json()
    name = data.get('name', '')
    folder = startup_folder(name)
    if not os.path.isdir(folder):
        os.makedirs(folder, exist_ok=True)
    import platform
    if platform.system() == 'Darwin':
        subprocess.Popen(['open', folder])
    return jsonify({'ok': True, 'path': folder})

@app.route('/api/update', methods=['POST'])
@login_required
def api_add_update():
    data = request.get_json()
    name = data.get('name', '')
    text = data.get('text', '')
    date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb.active
        header_row, col_map = find_excel_columns(ws)
        name_col = col_map.get('name')
        if not header_row or not name_col:
            return jsonify({'ok': False, 'error': 'Cannot locate header row'})
        row = find_company_row(ws, header_row, name_col, name)
        if not row:
            return jsonify({'ok': False, 'error': 'Company not found'})
        # Count existing filled note cells → append at n+1
        note_cols = get_note_col_indices(ws, header_row)
        filled = sum(1 for c in note_cols if ws.cell(row=row, column=c).value is not None)
        col = ensure_note_col(ws, header_row, filled + 1)
        ws.cell(row=row, column=col).value = format_note_cell(date, text)
        wb.save(EXCEL)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/update/delete', methods=['POST'])
@login_required
def api_delete_update():
    data = request.get_json()
    name = data.get('name', '')
    idx = data.get('index', -1)   # idx is 0-based newest-first (as displayed)
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb.active
        header_row, col_map = find_excel_columns(ws)
        name_col = col_map.get('name')
        if not header_row or not name_col:
            return jsonify({'ok': False, 'error': 'Cannot locate header row'})
        row = find_company_row(ws, header_row, name_col, name)
        if not row:
            return jsonify({'ok': False, 'error': 'Company not found'})

        note_cols = get_note_col_indices(ws, header_row)
        # Collect filled notes in oldest-first order (column order)
        filled = [(c, ws.cell(row=row, column=c).value)
                  for c in note_cols if ws.cell(row=row, column=c).value is not None]
        # idx is newest-first → convert to oldest-first index
        oldest_first_idx = len(filled) - 1 - idx
        if 0 <= oldest_first_idx < len(filled):
            filled.pop(oldest_first_idx)
        # Clear all note cells for this row, then re-write remaining
        for c in note_cols:
            ws.cell(row=row, column=c).value = None
        for i, (_, val) in enumerate(filled):
            col = ensure_note_col(ws, header_row, i + 1)
            ws.cell(row=row, column=col).value = val
        wb.save(EXCEL)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/status-options')
@login_required
def api_status_options():
    return jsonify(STATUS_OPTIONS)

@app.route('/api/debug')
@login_required
def api_debug():
    import traceback
    info = {
        'excel_path': EXCEL,
        'excel_exists': os.path.exists(EXCEL),
        'cwd': os.getcwd(),
        'base': BASE,
    }
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb.active
        headers = {c: ws.cell(row=1, column=c).value for c in range(1, 30) if ws.cell(row=1, column=c).value}
        info['headers'] = headers
        info['max_row'] = ws.max_row
    except Exception as e:
        info['openpyxl_error'] = str(e)
    try:
        rows = load_startups()
        info['startup_count'] = len(rows)
        if rows:
            info['first_startup'] = rows[0]
    except Exception as e:
        info['load_error'] = str(e)
        info['traceback'] = traceback.format_exc()
    return jsonify(info)

# ── Add startup ──────────────────────────────────────────────────────────────

@app.route('/api/startups/add', methods=['POST'])
@login_required
def api_add_startup():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'Company name is required'})
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb.active

        # Build header → column index map from row 1
        hmap = {}
        for c in range(1, 40):
            v = ws.cell(row=1, column=c).value
            if v:
                hmap[str(v).strip().lstrip('\ufeff')] = c

        name_col = hmap.get('Company name', 1)

        # Duplicate check
        for r in range(2, ws.max_row + 1):
            existing = ws.cell(row=r, column=name_col).value
            if existing and str(existing).strip().lower() == name.lower():
                return jsonify({'ok': False, 'error': f'"{name}" already exists in the database'})

        next_row = ws.max_row + 1

        def w(header, value):
            col = hmap.get(header)
            if col and value is not None:
                ws.cell(row=next_row, column=col).value = value

        w('Company name',      name)
        w('Category',          data.get('category'))
        w('Status',            data.get('status') or 'No contact')
        w('Country',           data.get('country'))
        w('Description',       data.get('description'))
        w('Status comment',    data.get('status_comment'))
        w('WBWSite',           data.get('website'))
        w('Added by',          data.get('added_by'))
        w('Key contact',       data.get('key_contact'))
        w('Key contact email', data.get('key_contact_email'))
        w('Relationship owner',data.get('relationship_owner'))
        w('Avoid',             data.get('avoid', 'No'))
        # EDF PV column header has trailing spaces in the file
        edf_col = next((c for h, c in hmap.items() if 'pv' in h.lower() and 'portfolio' in h.lower()), None)
        if edf_col:
            ws.cell(row=next_row, column=edf_col).value = data.get('edf_pv', 'No')
        w('NDA signed', data.get('nda', 'No'))

        # BU checkboxes
        bus = data.get('business_units', [])
        for bu in BU_CHECKBOX_COLS:
            col = hmap.get(bu)
            if col:
                ws.cell(row=next_row, column=col).value = (bu in bus)

        wb.save(EXCEL)
        os.makedirs(startup_folder(name), exist_ok=True)
        return jsonify({'ok': True, 'name': name})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ── Edit startup ─────────────────────────────────────────────────────────────

@app.route('/api/startups/edit', methods=['POST'])
@login_required
def api_edit_startup():
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'Company name is required'})
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb.active
        # Build full header map
        hmap = {}
        for c in range(1, 100):
            v = ws.cell(row=1, column=c).value
            if v:
                hmap[str(v).strip().lstrip('\ufeff')] = c
        last_col = max(hmap.values()) if hmap else 1
        # Ensure new columns exist
        for h in ['Key contact', 'Key contact email', 'Relationship owner', 'Avoid']:
            if h not in hmap:
                last_col += 1
                ws.cell(row=1, column=last_col).value = h
                hmap[h] = last_col
        name_col = hmap.get('Company name', 1)
        row = find_company_row(ws, 1, name_col, name)
        if not row:
            return jsonify({'ok': False, 'error': 'Company not found'})

        def w(header, value):
            col = hmap.get(header)
            if col:
                ws.cell(row=row, column=col).value = value if value not in (None, '') else None

        w('Category',          data.get('category'))
        w('Status',            data.get('status'))
        w('Country',           data.get('country'))
        w('Description',       data.get('description'))
        w('Status comment',    data.get('status_comment'))
        w('WBWSite',           data.get('website'))
        w('Added by',          data.get('added_by'))
        w('Key contact',       data.get('key_contact'))
        w('Key contact email', data.get('key_contact_email'))
        w('Relationship owner',data.get('relationship_owner'))
        w('Avoid',             data.get('avoid', 'No'))
        edf_col = next((c for h, c in hmap.items() if 'pv' in h.lower() and 'portfolio' in h.lower()), None)
        if edf_col:
            ws.cell(row=row, column=edf_col).value = data.get('edf_pv', 'No')
        w('NDA signed', data.get('nda', 'No'))
        # BU checkboxes
        bus = data.get('business_units', [])
        for bu in BU_CHECKBOX_COLS:
            col = hmap.get(bu)
            if col:
                ws.cell(row=row, column=col).value = (bu in bus)
        wb.save(EXCEL)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ── Bulk upload ───────────────────────────────────────────────────────────────

UPLOAD_COL_ALIASES = {
    'company name': 'Company name', 'company': 'Company name', 'name': 'Company name',
    'website': 'WBWSite', 'wbwsite': 'WBWSite', 'url': 'WBWSite',
    'category': 'Category', 'status': 'Status', 'country': 'Country',
    'description': 'Description', 'status comment': 'Status comment',
    'added by': 'Added by', 'edf pv portfolio': 'EDF PV portfolio', 'edf pv': 'EDF PV portfolio',
    'nda signed': 'NDA signed', 'nda': 'NDA signed',
    'key contact': 'Key contact', 'key contact name': 'Key contact',
    'key contact email': 'Key contact email',
    'relationship owner': 'Relationship owner', 'avoid': 'Avoid',
}

@app.route('/api/startups/bulk-upload', methods=['POST'])
@login_required
def api_bulk_upload():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file uploaded'})
    f = request.files['file']
    default_category = request.form.get('default_category', '').strip() or None
    try:
        upload_df = pd.read_excel(f, header=0)
        upload_df.columns = [str(c).strip() for c in upload_df.columns]
        # Map upload columns to canonical headers
        col_map = {}
        for col in upload_df.columns:
            canonical = UPLOAD_COL_ALIASES.get(col.lower())
            if canonical:
                col_map[col] = canonical
        upload_df = upload_df.rename(columns=col_map)
        if 'Company name' not in upload_df.columns:
            return jsonify({'ok': False, 'error': 'Could not find a "Company name" column in the uploaded file'})

        # Load existing names for duplicate check
        existing = load_startups()
        existing_names = {s['name'].lower() for s in existing}

        wb = openpyxl.load_workbook(EXCEL)
        ws = wb.active
        hmap = {}
        for c in range(1, 100):
            v = ws.cell(row=1, column=c).value
            if v:
                hmap[str(v).strip().lstrip('\ufeff')] = c
        last_col = max(hmap.values()) if hmap else 1
        for h in ['Key contact', 'Key contact email', 'Relationship owner', 'Avoid']:
            if h not in hmap:
                last_col += 1
                ws.cell(row=1, column=last_col).value = h
                hmap[h] = last_col

        added, skipped = 0, 0
        for _, row in upload_df.iterrows():
            name = str(row.get('Company name', '') or '').strip()
            if not name or name.lower() in ('nan', 'none'):
                continue
            if name.lower() in existing_names:
                skipped += 1
                continue
            next_row = ws.max_row + 1

            def w(header, value):
                col = hmap.get(header)
                if col and value is not None and str(value).lower() not in ('nan', 'none', ''):
                    ws.cell(row=next_row, column=col).value = value

            w('Company name',      name)
            w('Category',          row.get('Category') or default_category)
            w('Status',            row.get('Status') or 'No contact')
            w('Country',           row.get('Country'))
            w('Description',       row.get('Description'))
            w('Status comment',    row.get('Status comment'))
            w('WBWSite',           row.get('WBWSite'))
            w('Added by',          row.get('Added by'))
            w('Key contact',       row.get('Key contact'))
            w('Key contact email', row.get('Key contact email'))
            w('Relationship owner',row.get('Relationship owner'))
            w('Avoid',             row.get('Avoid') or 'No')
            edf_col = next((c for h, c in hmap.items() if 'pv' in h.lower() and 'portfolio' in h.lower()), None)
            if edf_col:
                ws.cell(row=next_row, column=edf_col).value = row.get('EDF PV portfolio') or 'No'
            w('NDA signed', row.get('NDA signed') or 'No')
            existing_names.add(name.lower())
            added += 1

        if added:
            wb.save(EXCEL)
        return jsonify({'ok': True, 'added': added, 'skipped': skipped})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ── Export collection ─────────────────────────────────────────────────────────

@app.route('/api/export')
@login_required
def api_export():
    col_name = request.args.get('collection', '').strip()
    startups = load_startups()
    if col_name:
        cols = load_collections()
        members = set(cols.get(col_name, []))
        startups = [s for s in startups if s['name'] in members]

    rows = []
    for s in startups:
        rows.append({
            'Company name':      s['name'],
            'Category':          s.get('category'),
            'Theme':             s.get('theme'),
            'Business units':    s.get('business_unit'),
            'Status':            s.get('status'),
            'Status comment':    s.get('status_comment'),
            'Country':           s.get('country'),
            'Website':           s.get('website'),
            'Description':       s.get('description'),
            'Added by':          s.get('added_by'),
            'EDF PV portfolio':  s.get('edf_pv'),
            'NDA signed':        s.get('nda'),
            'Key contact':       s.get('key_contact'),
            'Key contact email': s.get('key_contact_email'),
            'Relationship owner':s.get('relationship_owner'),
            'Avoid':             s.get('avoid'),
        })

    df = pd.DataFrame(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Startups')
    output.seek(0)
    filename = f"{col_name or 'all_startups'}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

# ── Collections ───────────────────────────────────────────────────────────────

def load_collections():
    if os.path.exists(COLLECTIONS_FILE):
        with open(COLLECTIONS_FILE) as f:
            return json.load(f)
    return {}

def save_collections(cols):
    with open(COLLECTIONS_FILE, 'w') as f:
        json.dump(cols, f, indent=2)

@app.route('/api/collections')
@login_required
def api_get_collections():
    return jsonify(load_collections())

@app.route('/api/collections/create', methods=['POST'])
@login_required
def api_create_collection():
    name = request.get_json().get('name', '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'Name required'})
    cols = load_collections()
    if name in cols:
        return jsonify({'ok': False, 'error': 'Already exists'})
    cols[name] = []
    save_collections(cols)
    return jsonify({'ok': True})

@app.route('/api/collections/add', methods=['POST'])
@login_required
def api_collection_add():
    data = request.get_json()
    col_name = data.get('collection', '')
    startup  = data.get('startup', '')
    cols = load_collections()
    if col_name not in cols:
        return jsonify({'ok': False, 'error': 'Collection not found'})
    if startup not in cols[col_name]:
        cols[col_name].append(startup)
    save_collections(cols)
    return jsonify({'ok': True})

@app.route('/api/collections/remove', methods=['POST'])
@login_required
def api_collection_remove():
    data = request.get_json()
    col_name = data.get('collection', '')
    startup  = data.get('startup', '')
    cols = load_collections()
    if col_name in cols and startup in cols[col_name]:
        cols[col_name].remove(startup)
    save_collections(cols)
    return jsonify({'ok': True})

@app.route('/api/collections/delete', methods=['POST'])
@login_required
def api_collection_delete():
    name = request.get_json().get('name', '')
    cols = load_collections()
    cols.pop(name, None)
    save_collections(cols)
    return jsonify({'ok': True})

# ── AI Tasks ─────────────────────────────────────────────────────────────────

_ai_tasks = {}   # task_id → {'status': 'running'|'done'|'error', 'result': ..., 'started': ...}

def _run_task_bg(task_id, fn):
    try:
        result = fn(EXCEL)
        _ai_tasks[task_id].update(status='done', result=result)
    except Exception as e:
        _ai_tasks[task_id].update(status='error', result={'error': str(e)})

@app.route('/api/ai/daily-augment', methods=['POST'])
@login_required
def api_ai_daily_augment():
    task_id = secrets.token_hex(8)
    _ai_tasks[task_id] = {'status': 'running', 'started': datetime.now().isoformat(), 'result': None}
    threading.Thread(target=_run_task_bg, args=(task_id, ai_tasks.daily_augment), daemon=True).start()
    return jsonify({'ok': True, 'task_id': task_id})

@app.route('/api/ai/monthly-update', methods=['POST'])
@login_required
def api_ai_monthly_update():
    task_id = secrets.token_hex(8)
    _ai_tasks[task_id] = {'status': 'running', 'started': datetime.now().isoformat(), 'result': None}
    threading.Thread(target=_run_task_bg, args=(task_id, ai_tasks.monthly_update), daemon=True).start()
    return jsonify({'ok': True, 'task_id': task_id})

@app.route('/api/ai/status/<task_id>')
@login_required
def api_ai_status(task_id):
    task = _ai_tasks.get(task_id)
    if not task:
        return jsonify({'ok': False, 'error': 'Unknown task'}), 404
    return jsonify({'ok': True, **task})

@app.route('/api/ai/find-similar', methods=['POST'])
@login_required
def api_ai_find_similar():
    name = request.get_json().get('name', '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'Name required'})
    task_id = secrets.token_hex(8)
    _ai_tasks[task_id] = {'status': 'running', 'started': datetime.now().isoformat(), 'result': None}
    def run():
        try:
            result = ai_tasks.find_similar_for_startup(name, EXCEL)
            _ai_tasks[task_id].update(status='done', result=result)
        except Exception as e:
            _ai_tasks[task_id].update(status='error', result={'error': str(e)})
    threading.Thread(target=run, daemon=True).start()
    return jsonify({'ok': True, 'task_id': task_id})

@app.route('/api/ai/fill-descriptions', methods=['POST'])
@login_required
def api_ai_fill_descriptions():
    task_id = secrets.token_hex(8)
    _ai_tasks[task_id] = {'status': 'running', 'started': datetime.now().isoformat(), 'result': None}
    threading.Thread(target=_run_task_bg, args=(task_id, ai_tasks.fill_missing_descriptions), daemon=True).start()
    return jsonify({'ok': True, 'task_id': task_id})

@app.route('/api/ai/fill-websites', methods=['POST'])
@login_required
def api_ai_fill_websites():
    task_id = secrets.token_hex(8)
    _ai_tasks[task_id] = {'status': 'running', 'started': datetime.now().isoformat(), 'result': None}
    threading.Thread(target=_run_task_bg, args=(task_id, ai_tasks.fill_missing_websites), daemon=True).start()
    return jsonify({'ok': True, 'task_id': task_id})

@app.route('/api/ai/log')
@login_required
def api_ai_log():
    return jsonify(ai_tasks.get_log())

# ── Health check (no auth — required by AWS ALB / ECS / App Runner) ──────────

@app.route('/health')
def health():
    return jsonify({'ok': True})

# ── Boot ─────────────────────────────────────────────────────────────────────

def ensure_new_columns():
    """Add new field column headers to the spreadsheet if they don't exist yet."""
    try:
        wb = openpyxl.load_workbook(EXCEL)
        ws = wb.active
        hmap = {}
        for c in range(1, 100):
            v = ws.cell(row=1, column=c).value
            if v:
                hmap[str(v).strip().lstrip('\ufeff')] = c
        last_col = max(hmap.values()) if hmap else 1
        needed = ['Key contact', 'Key contact email', 'Relationship owner', 'Avoid']
        changed = False
        for h in needed:
            if h not in hmap:
                last_col += 1
                ws.cell(row=1, column=last_col).value = h
                changed = True
        if changed:
            wb.save(EXCEL)
    except Exception as e:
        print(f'[ensure_columns] {e}')

ensure_new_columns()
migrate_notes_once()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=True, port=port)
